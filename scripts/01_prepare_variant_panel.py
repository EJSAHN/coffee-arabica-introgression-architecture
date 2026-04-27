#!/usr/bin/env python3
"""
Build analysis-ready subgenome SNP panels from public Coffea arabica resources.

The script reads accession metadata, subgenome-resolved VCF files, and the
syntenic-alignment archive distributed with the public arabica population-genomic
resource. It summarizes metadata, VCF headers, sample-level quality metrics,
contig-level variant counts, deterministic SNP subsets, PCA scores,
identity-by-state-like similarity matrices, and group-level allele-frequency
contrasts.
"""

from __future__ import annotations

import argparse
import collections
import gzip
import io
import json
import math
import os
import random
import re
import statistics
import sys
import tarfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


EXPECTED_FILES = {
    "metadata": "Accession_info.xlsx",
    "sgC_vcf": "Arabica_sgC.TIP.BB.vcf.gz",
    "sgE_vcf": "Arabica_sgE.TIP.BB.vcf.gz",
    "synteny": "Coffea_syntenic_alignments.tar.gz",
    "readme": "README.md",
}


@dataclass
class VariantRecord:
    contig: str
    pos: int
    variant_id: str
    ref: str
    alt: str
    qual: str
    filter_value: str
    call_rate: float
    maf: float
    dosages: List[float]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run a manuscript-first exploratory analysis of the Coffea introgression dataset."
    )
    parser.add_argument(
        "--project-dir",
        required=True,
        help="Project directory containing the five expected input files."
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory for Excel outputs. Defaults to <project-dir>/analysis_outputs."
    )
    parser.add_argument(
        "--max-analysis-variants-per-vcf",
        type=int,
        default=12000,
        help="Maximum number of filtered variants retained per VCF for PCA, IBS and group contrasts."
    )
    parser.add_argument(
        "--min-site-call-rate",
        type=float,
        default=0.80,
        help="Minimum site call rate required for inclusion in the analysis-ready variant panel."
    )
    parser.add_argument(
        "--maf-threshold",
        type=float,
        default=0.05,
        help="Minimum minor allele frequency required for inclusion in the analysis-ready variant panel."
    )
    parser.add_argument(
        "--min-group-size",
        type=int,
        default=2,
        help="Minimum number of accessions required for a group to be included in allele-frequency contrasts."
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=250000,
        help="Write a progress message every N parsed variants."
    )
    parser.add_argument(
        "--random-seed",
        type=int,
        default=20250416,
        help="Random seed used for deterministic reservoir sampling."
    )
    return parser.parse_args()


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_species(species: str) -> str:
    s = normalize_text(species).lower()
    if s == "coffea arabica":
        return "Coffea arabica"
    if s == "coffea canephora":
        return "Coffea canephora"
    if s == "coffea eugenioides":
        return "Coffea eugenioides"
    return normalize_text(species)


def normalize_variety(variety: str) -> str:
    v = normalize_text(variety).lower()
    if "introgress" in v:
        return "Introgressed"
    if "cultivated" in v:
        return "Cultivated"
    if "wild" in v:
        return "Wild"
    return normalize_text(variety)


def infer_analysis_group(species_name: str, variety_name: str) -> str:
    species = normalize_species(species_name)
    variety = normalize_variety(variety_name)
    if species == "Coffea arabica":
        if variety == "Introgressed":
            return "arabica_introgressed"
        if variety == "Cultivated":
            return "arabica_cultivated"
        if variety == "Wild":
            return "arabica_wild"
        return "arabica_other"
    if species == "Coffea canephora":
        return "canephora"
    if species == "Coffea eugenioides":
        return "eugenioides"
    return "other"


def locate_inputs(project_dir: Path) -> Dict[str, Path]:
    located = {}
    for key, filename in EXPECTED_FILES.items():
        candidate = project_dir / filename
        if not candidate.exists():
            raise FileNotFoundError(
                f"Expected input file not found: {candidate}"
            )
        located[key] = candidate
    return located


def load_and_clean_metadata(metadata_path: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    raw_df = pd.read_excel(metadata_path)
    clean_df = raw_df.copy()

    rename_map = {
        "Seq.ID": "seq_id",
        "Collection location ": "collection_location",
        "species_name": "species_name",
        "variety(s)": "variety",
        "location_code(s)": "location_code",
        "country_of_origin(s)": "country_of_origin",
        "ploidy_level(s)": "ploidy_level",
        "Genome size (Gb)": "genome_size_gb",
        "genome_structure(s)": "genome_structure",
        "donor_institute(s)": "donor_institute",
        "notes(s)": "notes",
        "Full ref": "full_reference",
    }
    clean_df = clean_df.rename(columns=rename_map)

    clean_df["seq_id"] = clean_df["seq_id"].map(normalize_text)
    clean_df["accession_name"] = clean_df["accession_name"].map(normalize_text)
    clean_df["species_name"] = clean_df["species_name"].map(normalize_species)
    clean_df["variety"] = clean_df["variety"].map(normalize_variety)
    clean_df["analysis_group"] = clean_df.apply(
        lambda row: infer_analysis_group(row.get("species_name", ""), row.get("variety", "")),
        axis=1
    )
    clean_df["has_latitude"] = clean_df["Latitude"].notna()
    clean_df["has_longitude"] = clean_df["Longitude"].notna()
    clean_df["has_altitude"] = clean_df["altitude"].notna()

    clean_df["seq_id"] = clean_df["seq_id"].astype(str)
    clean_df["order"] = pd.to_numeric(clean_df["order"], errors="coerce")

    return raw_df, clean_df


def inventory_synteny_tar(tar_path: Path) -> pd.DataFrame:
    rows = []
    with tarfile.open(tar_path, "r:gz") as tf:
        for member in tf.getmembers():
            if member.isdir():
                continue
            is_macos_sidecar = Path(member.name).name.startswith("._")
            clean_name = Path(member.name).name
            clean_name_no_sidecar = clean_name[2:] if is_macos_sidecar else clean_name

            parts = clean_name_no_sidecar.split(".")
            pair_token = parts[0] if parts else clean_name_no_sidecar
            pair_match = re.match(r"^(\d+)_(\d+)$", pair_token)

            genome_a = pair_match.group(1) if pair_match else ""
            genome_b = pair_match.group(2) if pair_match else ""

            if "tandems" in clean_name_no_sidecar:
                file_category = "tandem_duplicates"
            elif "aligncoords" in clean_name_no_sidecar:
                file_category = "syntenic_alignment"
            else:
                file_category = "other"

            rows.append(
                {
                    "member_name": clean_name,
                    "member_size_bytes": member.size,
                    "is_macos_sidecar": is_macos_sidecar,
                    "clean_member_name": clean_name_no_sidecar,
                    "genome_a_id": genome_a,
                    "genome_b_id": genome_b,
                    "file_category": file_category,
                }
            )
    return pd.DataFrame(rows)


def parse_contig_header_line(line: str) -> Tuple[str, Optional[int]]:
    # Example:
    # ##contig=<ID=chr_A_sg_C_(Scaffold_18;HRSCAF_19),length=35656778>
    contig_match = re.search(r"ID=([^,>]+)", line)
    length_match = re.search(r"length=(\d+)", line)
    contig_id = contig_match.group(1) if contig_match else ""
    length = int(length_match.group(1)) if length_match else None
    return contig_id, length


def extract_header_metadata(header_lines: List[str], sample_names: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    summary_rows = []
    contig_rows = []

    fileformat = ""
    genotypegvcfs_version = ""
    haplotypecaller_version = ""
    genotypegvcfs_ploidy = ""
    haplotypecaller_ploidy = ""
    combined_reference = ""
    subgenome_reference = ""

    for line in header_lines:
        if line.startswith("##fileformat="):
            fileformat = line.split("=", 1)[1].strip()
        elif line.startswith("##GATKCommandLine.GenotypeGVCFs"):
            version_match = re.search(r"Version=([^,>]+)", line)
            ploidy_match = re.search(r"sample_ploidy=([0-9]+)", line)
            reference_match = re.search(r"reference_sequence=([^ ]+)", line)
            genotypegvcfs_version = version_match.group(1) if version_match else ""
            genotypegvcfs_ploidy = ploidy_match.group(1) if ploidy_match else ""
            combined_reference = reference_match.group(1) if reference_match else ""
        elif line.startswith("##GATKCommandLine.HaplotypeCaller"):
            version_match = re.search(r"Version=([^,>]+)", line)
            ploidy_match = re.search(r"sample_ploidy=([0-9]+)", line)
            reference_match = re.search(r"reference_sequence=([^ ]+)", line)
            haplotypecaller_version = version_match.group(1) if version_match else ""
            haplotypecaller_ploidy = ploidy_match.group(1) if ploidy_match else ""
            subgenome_reference = reference_match.group(1) if reference_match else ""
        elif line.startswith("##contig="):
            contig_id, length = parse_contig_header_line(line)
            contig_rows.append(
                {
                    "contig_id": contig_id,
                    "contig_length": length,
                }
            )

    summary_rows.extend(
        [
            {"field": "fileformat", "value": fileformat},
            {"field": "sample_count", "value": len(sample_names)},
            {"field": "contig_count", "value": len(contig_rows)},
            {"field": "GenotypeGVCFs_version", "value": genotypegvcfs_version},
            {"field": "HaplotypeCaller_version", "value": haplotypecaller_version},
            {"field": "GenotypeGVCFs_sample_ploidy", "value": genotypegvcfs_ploidy},
            {"field": "HaplotypeCaller_sample_ploidy", "value": haplotypecaller_ploidy},
            {"field": "combined_reference_path", "value": combined_reference},
            {"field": "subgenome_reference_path", "value": subgenome_reference},
        ]
    )

    return pd.DataFrame(summary_rows), pd.DataFrame(contig_rows)


def parse_genotype(sample_value: str, gt_index: int, dp_index: Optional[int]) -> Tuple[float, Optional[int], bool]:
    if sample_value in {".", "./.", ".|.", ""}:
        return np.nan, None, True

    fields = sample_value.split(":")
    gt = fields[gt_index] if gt_index < len(fields) else "."
    dp = None
    if dp_index is not None and dp_index < len(fields):
        try:
            dp = int(fields[dp_index])
        except Exception:
            dp = None

    if gt in {".", "./.", ".|."}:
        return np.nan, dp, True

    allele_tokens = re.split(r"[\/|]", gt)
    if any(token == "." or token == "" for token in allele_tokens):
        return np.nan, dp, True

    try:
        allele_ints = [int(token) for token in allele_tokens]
    except ValueError:
        return np.nan, dp, True

    return float(sum(allele_ints)), dp, False


def reservoir_replace(reservoir: List[VariantRecord], record: VariantRecord, seen_count: int, max_size: int, rng: random.Random) -> None:
    if len(reservoir) < max_size:
        reservoir.append(record)
        return
    replacement_index = rng.randint(0, seen_count - 1)
    if replacement_index < max_size:
        reservoir[replacement_index] = record


def compute_pca(sample_names: List[str], variant_records: List[VariantRecord], n_components: int = 5) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not variant_records:
        return pd.DataFrame(), pd.DataFrame()

    dosage_matrix = np.array([record.dosages for record in variant_records], dtype=float)  # variants x samples
    if dosage_matrix.size == 0:
        return pd.DataFrame(), pd.DataFrame()

    # Impute missing dosage with per-variant mean.
    row_means = np.nanmean(dosage_matrix, axis=1)
    row_means = np.where(np.isnan(row_means), 0.0, row_means)
    nan_mask = np.isnan(dosage_matrix)
    dosage_matrix[nan_mask] = np.take(row_means, np.where(nan_mask)[0])

    # Transpose to samples x variants
    X = dosage_matrix.T

    # Standardize by variant.
    means = X.mean(axis=0)
    stds = X.std(axis=0, ddof=1)
    stds = np.where((stds == 0) | np.isnan(stds), 1.0, stds)
    X_std = (X - means) / stds

    # Singular value decomposition
    U, S, Vt = np.linalg.svd(X_std, full_matrices=False)
    max_components = min(n_components, U.shape[1])
    scores = U[:, :max_components] * S[:max_components]
    explained_variance = (S ** 2) / max(1, (X_std.shape[0] - 1))
    total_variance = explained_variance.sum()
    explained_ratio = explained_variance / total_variance if total_variance > 0 else np.zeros_like(explained_variance)

    score_df = pd.DataFrame({"sample_id": sample_names})
    for i in range(max_components):
        score_df[f"PC{i+1}"] = scores[:, i]
        score_df[f"PC{i+1}_explained_variance_ratio"] = explained_ratio[i]

    variance_df = pd.DataFrame(
        {
            "component": [f"PC{i+1}" for i in range(max_components)],
            "explained_variance_ratio": explained_ratio[:max_components],
        }
    )
    return score_df, variance_df


def compute_ibs_matrix(sample_names: List[str], variant_records: List[VariantRecord]) -> pd.DataFrame:
    if not variant_records:
        return pd.DataFrame()

    n_samples = len(sample_names)
    sim_sum = np.zeros((n_samples, n_samples), dtype=float)
    sim_n = np.zeros((n_samples, n_samples), dtype=float)

    for record in variant_records:
        dosages = np.array(record.dosages, dtype=float)
        valid = ~np.isnan(dosages)
        for i in range(n_samples):
            if not valid[i]:
                continue
            for j in range(i, n_samples):
                if not valid[j]:
                    continue
                similarity = 1.0 - abs(dosages[i] - dosages[j]) / 2.0
                sim_sum[i, j] += similarity
                sim_n[i, j] += 1.0
                if i != j:
                    sim_sum[j, i] += similarity
                    sim_n[j, i] += 1.0

    with np.errstate(divide="ignore", invalid="ignore"):
        sim = np.divide(sim_sum, sim_n, out=np.full_like(sim_sum, np.nan), where=sim_n > 0)

    ibs_df = pd.DataFrame(sim, index=sample_names, columns=sample_names)
    ibs_df.index.name = "sample_id"
    return ibs_df


def compute_group_summary(sample_qc_df: pd.DataFrame, metadata_subset_df: pd.DataFrame) -> pd.DataFrame:
    if sample_qc_df.empty:
        return pd.DataFrame()

    merged = sample_qc_df.merge(metadata_subset_df, how="left", left_on="sample_id", right_on="seq_id")
    numeric_cols = [
        "call_rate_all_sites",
        "call_rate_biallelic_snps",
        "heterozygosity_rate_biallelic_snps",
        "non_reference_rate_biallelic_snps",
        "mean_depth",
    ]
    available_numeric = [col for col in numeric_cols if col in merged.columns]

    group_df = (
        merged.groupby("analysis_group", dropna=False)
        .agg(
            n_samples=("sample_id", "size"),
            **{f"mean_{col}": (col, "mean") for col in available_numeric}
        )
        .reset_index()
        .sort_values(["n_samples", "analysis_group"], ascending=[False, True])
    )
    return group_df


def compute_group_contrasts(
    variant_records: List[VariantRecord],
    sample_names: List[str],
    metadata_subset_df: pd.DataFrame,
    min_group_size: int,
    top_n_per_contrast: int = 250
) -> pd.DataFrame:
    if not variant_records:
        return pd.DataFrame()

    meta = metadata_subset_df.copy()
    meta["seq_id"] = meta["seq_id"].astype(str)
    meta = meta.set_index("seq_id", drop=False)

    group_to_indices = collections.OrderedDict()
    for idx, sample_id in enumerate(sample_names):
        if sample_id in meta.index:
            group = meta.loc[sample_id, "analysis_group"]
        else:
            group = "unmatched"
        group_to_indices.setdefault(group, []).append(idx)

    eligible_groups = {
        group: indices
        for group, indices in group_to_indices.items()
        if len(indices) >= min_group_size
    }

    default_contrasts = [
        ("arabica_cultivated", "arabica_introgressed"),
        ("arabica_cultivated", "arabica_wild"),
        ("arabica_cultivated", "canephora"),
        ("arabica_cultivated", "eugenioides"),
        ("arabica_introgressed", "canephora"),
        ("arabica_introgressed", "eugenioides"),
        ("arabica_wild", "canephora"),
        ("arabica_wild", "eugenioides"),
    ]

    contrast_rows = []
    for record in variant_records:
        dosages = np.array(record.dosages, dtype=float)
        variant_group_af = {}
        for group, indices in eligible_groups.items():
            group_dosages = dosages[indices]
            called = group_dosages[~np.isnan(group_dosages)]
            if called.size == 0:
                continue
            variant_group_af[group] = float(called.sum() / (2.0 * called.size))

        for g1, g2 in default_contrasts:
            if g1 not in variant_group_af or g2 not in variant_group_af:
                continue
            af1 = variant_group_af[g1]
            af2 = variant_group_af[g2]
            contrast_rows.append(
                {
                    "contrast": f"{g1}__vs__{g2}",
                    "contig": record.contig,
                    "pos": record.pos,
                    "variant_id": record.variant_id,
                    "ref": record.ref,
                    "alt": record.alt,
                    "maf_total": record.maf,
                    "call_rate_total": record.call_rate,
                    "group1": g1,
                    "group2": g2,
                    "group1_af": af1,
                    "group2_af": af2,
                    "abs_delta_af": abs(af1 - af2),
                }
            )

    if not contrast_rows:
        return pd.DataFrame()

    contrast_df = pd.DataFrame(contrast_rows)
    contrast_df = contrast_df.sort_values(
        ["contrast", "abs_delta_af", "maf_total"],
        ascending=[True, False, False]
    )
    top_df = (
        contrast_df.groupby("contrast", as_index=False, sort=False)
        .head(top_n_per_contrast)
        .reset_index(drop=True)
    )
    return top_df


def parse_vcf_stream(
    vcf_path: Path,
    metadata_df: pd.DataFrame,
    max_analysis_variants: int,
    min_site_call_rate: float,
    maf_threshold: float,
    progress_every: int,
    random_seed: int,
    min_group_size: int,
) -> Dict[str, object]:
    print(f"Starting VCF parse: {vcf_path.name}")
    rng = random.Random(random_seed)

    header_lines: List[str] = []
    sample_names: List[str] = []
    format_cache: Dict[str, Tuple[Optional[int], Optional[int]]] = {}
    contig_summary = collections.defaultdict(
        lambda: {
            "variant_count": 0,
            "biallelic_snp_count": 0,
            "analysis_ready_variant_count": 0,
        }
    )

    global_counts = collections.Counter()
    filter_counts = collections.Counter()

    sample_stats = None
    sampled_variants: List[VariantRecord] = []
    analysis_ready_seen = 0

    with gzip.open(vcf_path, "rt", encoding="utf-8", errors="replace") as handle:
        for raw_line in handle:
            if raw_line.startswith("##"):
                header_lines.append(raw_line.rstrip("\n"))
                continue

            if raw_line.startswith("#CHROM"):
                columns = raw_line.rstrip("\n").split("\t")
                sample_names = columns[9:]
                header_lines.append(raw_line.rstrip("\n"))

                sample_stats = {
                    sample_id: {
                        "total_sites": 0,
                        "called_sites": 0,
                        "missing_sites": 0,
                        "biallelic_snp_sites": 0,
                        "called_biallelic_snp_sites": 0,
                        "het_biallelic_snp_sites": 0,
                        "nonref_biallelic_snp_sites": 0,
                        "total_depth_observations": 0,
                        "sum_depth": 0,
                    }
                    for sample_id in sample_names
                }
                continue

            if not sample_names:
                raise RuntimeError("VCF header is malformed: sample names were not found before data lines.")

            fields = raw_line.rstrip("\n").split("\t")
            if len(fields) < 10:
                continue

            chrom, pos, variant_id, ref, alt, qual, filt, info, fmt = fields[:9]
            sample_values = fields[9:]

            global_counts["variant_lines_total"] += 1
            contig_summary[chrom]["variant_count"] += 1
            filter_counts[filt] += 1

            alt_alleles = alt.split(",")
            is_biallelic = len(alt_alleles) == 1
            is_snp = len(ref) == 1 and all(len(a) == 1 and a != "*" for a in alt_alleles)

            if is_snp:
                global_counts["snp_variant_lines"] += 1
            else:
                global_counts["non_snp_variant_lines"] += 1

            if is_biallelic and is_snp:
                global_counts["biallelic_snp_variant_lines"] += 1
                contig_summary[chrom]["biallelic_snp_count"] += 1
            else:
                global_counts["non_biallelic_or_non_snp_variant_lines"] += 1

            if fmt not in format_cache:
                format_fields = fmt.split(":")
                gt_index = format_fields.index("GT") if "GT" in format_fields else None
                dp_index = format_fields.index("DP") if "DP" in format_fields else None
                format_cache[fmt] = (gt_index, dp_index)

            gt_index, dp_index = format_cache[fmt]
            if gt_index is None:
                raise RuntimeError("VCF FORMAT field does not contain GT.")

            dosages = []
            called_count = 0
            allele_number = 0
            alt_allele_count = 0
            het_count = 0

            for sample_id, sample_value in zip(sample_names, sample_values):
                dosage, depth, missing = parse_genotype(sample_value, gt_index, dp_index)

                sample_stats[sample_id]["total_sites"] += 1
                if missing:
                    sample_stats[sample_id]["missing_sites"] += 1
                else:
                    sample_stats[sample_id]["called_sites"] += 1
                    called_count += 1
                    if depth is not None:
                        sample_stats[sample_id]["total_depth_observations"] += 1
                        sample_stats[sample_id]["sum_depth"] += depth

                dosages.append(dosage)

                if is_biallelic and is_snp:
                    sample_stats[sample_id]["biallelic_snp_sites"] += 1
                    if not missing:
                        sample_stats[sample_id]["called_biallelic_snp_sites"] += 1
                        allele_number += 2
                        alt_allele_count += int(dosage)
                        if dosage == 1:
                            het_count += 1
                            sample_stats[sample_id]["het_biallelic_snp_sites"] += 1
                        if dosage > 0:
                            sample_stats[sample_id]["nonref_biallelic_snp_sites"] += 1

            if global_counts["variant_lines_total"] % progress_every == 0:
                print(
                    f"Parsed {global_counts['variant_lines_total']:,} variants from {vcf_path.name}"
                )

            if not (is_biallelic and is_snp):
                continue

            site_call_rate = called_count / max(1, len(sample_names))
            alt_af = (alt_allele_count / allele_number) if allele_number > 0 else np.nan
            maf = min(alt_af, 1.0 - alt_af) if not np.isnan(alt_af) else np.nan

            if site_call_rate >= min_site_call_rate and not np.isnan(maf) and maf >= maf_threshold:
                analysis_ready_seen += 1
                contig_summary[chrom]["analysis_ready_variant_count"] += 1
                record = VariantRecord(
                    contig=chrom,
                    pos=int(pos),
                    variant_id=variant_id if variant_id != "." else "",
                    ref=ref,
                    alt=alt,
                    qual=qual,
                    filter_value=filt,
                    call_rate=site_call_rate,
                    maf=maf,
                    dosages=dosages,
                )
                reservoir_replace(sampled_variants, record, analysis_ready_seen, max_analysis_variants, rng)

    header_summary_df, header_contig_df = extract_header_metadata(header_lines, sample_names)

    filter_df = (
        pd.DataFrame(
            [{"filter_value": k, "count": v} for k, v in filter_counts.items()]
        )
        .sort_values("count", ascending=False)
        .reset_index(drop=True)
    )

    variant_summary_rows = [
        {"metric": "variant_lines_total", "value": global_counts.get("variant_lines_total", 0)},
        {"metric": "snp_variant_lines", "value": global_counts.get("snp_variant_lines", 0)},
        {"metric": "non_snp_variant_lines", "value": global_counts.get("non_snp_variant_lines", 0)},
        {"metric": "biallelic_snp_variant_lines", "value": global_counts.get("biallelic_snp_variant_lines", 0)},
        {"metric": "non_biallelic_or_non_snp_variant_lines", "value": global_counts.get("non_biallelic_or_non_snp_variant_lines", 0)},
        {"metric": "analysis_ready_variant_lines_before_sampling", "value": analysis_ready_seen},
        {"metric": "analysis_ready_variant_lines_retained", "value": len(sampled_variants)},
        {"metric": "min_site_call_rate", "value": min_site_call_rate},
        {"metric": "maf_threshold", "value": maf_threshold},
    ]
    variant_summary_df = pd.DataFrame(variant_summary_rows)

    contig_df = pd.DataFrame(
        [
            {"contig_id": k, **v}
            for k, v in contig_summary.items()
        ]
    ).sort_values(["analysis_ready_variant_count", "biallelic_snp_count", "variant_count"], ascending=False)

    sample_qc_rows = []
    for sample_id, stats in sample_stats.items():
        total_sites = stats["total_sites"]
        biallelic_sites = stats["biallelic_snp_sites"]
        called_sites = stats["called_sites"]
        called_biallelic_sites = stats["called_biallelic_snp_sites"]
        mean_depth = (
            stats["sum_depth"] / stats["total_depth_observations"]
            if stats["total_depth_observations"] > 0
            else np.nan
        )
        sample_qc_rows.append(
            {
                "sample_id": sample_id,
                "total_sites": total_sites,
                "called_sites": called_sites,
                "missing_sites": stats["missing_sites"],
                "call_rate_all_sites": called_sites / total_sites if total_sites > 0 else np.nan,
                "biallelic_snp_sites": biallelic_sites,
                "called_biallelic_snp_sites": called_biallelic_sites,
                "call_rate_biallelic_snps": called_biallelic_sites / biallelic_sites if biallelic_sites > 0 else np.nan,
                "het_biallelic_snp_sites": stats["het_biallelic_snp_sites"],
                "heterozygosity_rate_biallelic_snps": stats["het_biallelic_snp_sites"] / called_biallelic_sites if called_biallelic_sites > 0 else np.nan,
                "nonref_biallelic_snp_sites": stats["nonref_biallelic_snp_sites"],
                "non_reference_rate_biallelic_snps": stats["nonref_biallelic_snp_sites"] / called_biallelic_sites if called_biallelic_sites > 0 else np.nan,
                "mean_depth": mean_depth,
            }
        )
    sample_qc_df = pd.DataFrame(sample_qc_rows).sort_values("sample_id").reset_index(drop=True)

    metadata_subset = metadata_df[metadata_df["seq_id"].isin(sample_names)].copy()
    matched_ids = set(metadata_subset["seq_id"].tolist())
    sample_match_df = pd.DataFrame(
        {
            "sample_id": sample_names,
            "has_metadata_match": [sample_id in matched_ids for sample_id in sample_names],
        }
    ).merge(
        metadata_subset,
        how="left",
        left_on="sample_id",
        right_on="seq_id"
    )

    pca_scores_df, pca_variance_df = compute_pca(sample_names, sampled_variants, n_components=5)
    if not pca_scores_df.empty:
        pca_scores_df = pca_scores_df.merge(
            metadata_subset[
                [
                    "seq_id",
                    "accession_name",
                    "species_name",
                    "variety",
                    "analysis_group",
                    "genome_structure",
                    "country_of_origin",
                ]
            ],
            how="left",
            left_on="sample_id",
            right_on="seq_id",
        )

    ibs_df = compute_ibs_matrix(sample_names, sampled_variants)
    group_summary_df = compute_group_summary(sample_qc_df, metadata_subset)
    contrast_df = compute_group_contrasts(
        sampled_variants,
        sample_names,
        metadata_subset,
        min_group_size=min_group_size,
        top_n_per_contrast=250,
    )

    analysis_variant_df = pd.DataFrame(
        [
            {
                "contig": vr.contig,
                "pos": vr.pos,
                "variant_id": vr.variant_id,
                "ref": vr.ref,
                "alt": vr.alt,
                "qual": vr.qual,
                "filter_value": vr.filter_value,
                "call_rate": vr.call_rate,
                "maf": vr.maf,
            }
            for vr in sampled_variants
        ]
    ).sort_values(["contig", "pos"]).reset_index(drop=True)

    return {
        "sample_names": sample_names,
        "header_summary_df": header_summary_df,
        "header_contig_df": header_contig_df,
        "variant_summary_df": variant_summary_df,
        "filter_df": filter_df,
        "contig_df": contig_df,
        "sample_qc_df": sample_qc_df,
        "sample_match_df": sample_match_df,
        "pca_scores_df": pca_scores_df,
        "pca_variance_df": pca_variance_df,
        "ibs_df": ibs_df,
        "group_summary_df": group_summary_df,
        "contrast_df": contrast_df,
        "analysis_variant_df": analysis_variant_df,
    }


def build_input_inventory(file_map: Dict[str, Path]) -> pd.DataFrame:
    rows = []
    for key, path in file_map.items():
        rows.append(
            {
                "input_key": key,
                "filename": path.name,
                "absolute_path": str(path.resolve()),
                "file_size_bytes": path.stat().st_size,
                "file_size_mb": round(path.stat().st_size / (1024 ** 2), 3),
            }
        )
    return pd.DataFrame(rows)


def build_dataset_summary(
    clean_meta_df: pd.DataFrame,
    sgc_result: Dict[str, object],
    sge_result: Dict[str, object],
    synteny_df: pd.DataFrame,
) -> pd.DataFrame:
    rows = []

    rows.append({"metric": "metadata_rows", "value": len(clean_meta_df)})
    rows.append({"metric": "metadata_unique_seq_id", "value": clean_meta_df["seq_id"].nunique()})
    rows.append({"metric": "metadata_unique_accession_name", "value": clean_meta_df["accession_name"].nunique()})
    rows.append({"metric": "metadata_species_count", "value": clean_meta_df["species_name"].nunique()})
    rows.append({"metric": "metadata_analysis_group_count", "value": clean_meta_df["analysis_group"].nunique()})
    rows.append({"metric": "sgC_sample_count", "value": len(sgc_result["sample_names"])})
    rows.append({"metric": "sgE_sample_count", "value": len(sge_result["sample_names"])})
    rows.append({"metric": "sgC_variant_lines_total", "value": int(sgc_result["variant_summary_df"].loc[sgc_result["variant_summary_df"]["metric"] == "variant_lines_total", "value"].iloc[0])})
    rows.append({"metric": "sgE_variant_lines_total", "value": int(sge_result["variant_summary_df"].loc[sge_result["variant_summary_df"]["metric"] == "variant_lines_total", "value"].iloc[0])})
    rows.append({"metric": "sgC_analysis_variants_retained", "value": int(sgc_result["variant_summary_df"].loc[sgc_result["variant_summary_df"]["metric"] == "analysis_ready_variant_lines_retained", "value"].iloc[0])})
    rows.append({"metric": "sgE_analysis_variants_retained", "value": int(sge_result["variant_summary_df"].loc[sge_result["variant_summary_df"]["metric"] == "analysis_ready_variant_lines_retained", "value"].iloc[0])})
    rows.append({"metric": "synteny_tar_members_total", "value": len(synteny_df)})
    rows.append({"metric": "synteny_tar_members_non_sidecar", "value": int((~synteny_df["is_macos_sidecar"]).sum())})
    return pd.DataFrame(rows)


def build_sample_overlap(sgc_result: Dict[str, object], sge_result: Dict[str, object]) -> pd.DataFrame:
    sgc_samples = set(sgc_result["sample_names"])
    sge_samples = set(sge_result["sample_names"])

    all_samples = sorted(sgc_samples | sge_samples)
    rows = []
    for sample_id in all_samples:
        rows.append(
            {
                "sample_id": sample_id,
                "present_in_sgC": sample_id in sgc_samples,
                "present_in_sgE": sample_id in sge_samples,
                "present_in_both": (sample_id in sgc_samples) and (sample_id in sge_samples),
            }
        )
    return pd.DataFrame(rows)


def build_subgenome_asymmetry(
    clean_meta_df: pd.DataFrame,
    sgc_result: Dict[str, object],
    sge_result: Dict[str, object],
) -> pd.DataFrame:
    sgc_qc = sgc_result["sample_qc_df"].copy()
    sge_qc = sge_result["sample_qc_df"].copy()

    sgc_qc = sgc_qc.add_prefix("sgC_").rename(columns={"sgC_sample_id": "sample_id"})
    sge_qc = sge_qc.add_prefix("sgE_").rename(columns={"sgE_sample_id": "sample_id"})

    merged = sgc_qc.merge(sge_qc, how="outer", on="sample_id")
    merged = merged.merge(
        clean_meta_df[
            [
                "seq_id",
                "accession_name",
                "species_name",
                "variety",
                "analysis_group",
                "country_of_origin",
                "genome_structure",
            ]
        ],
        how="left",
        left_on="sample_id",
        right_on="seq_id",
    )

    if "sgC_heterozygosity_rate_biallelic_snps" in merged.columns and "sgE_heterozygosity_rate_biallelic_snps" in merged.columns:
        merged["delta_heterozygosity_sgC_minus_sgE"] = (
            merged["sgC_heterozygosity_rate_biallelic_snps"] - merged["sgE_heterozygosity_rate_biallelic_snps"]
        )
    if "sgC_call_rate_biallelic_snps" in merged.columns and "sgE_call_rate_biallelic_snps" in merged.columns:
        merged["delta_call_rate_sgC_minus_sgE"] = (
            merged["sgC_call_rate_biallelic_snps"] - merged["sgE_call_rate_biallelic_snps"]
        )
    if "sgC_non_reference_rate_biallelic_snps" in merged.columns and "sgE_non_reference_rate_biallelic_snps" in merged.columns:
        merged["delta_non_reference_rate_sgC_minus_sgE"] = (
            merged["sgC_non_reference_rate_biallelic_snps"] - merged["sgE_non_reference_rate_biallelic_snps"]
        )

    merged = merged.sort_values(["analysis_group", "sample_id"]).reset_index(drop=True)
    return merged


def build_group_summary(clean_meta_df: pd.DataFrame) -> pd.DataFrame:
    group_df = (
        clean_meta_df.groupby(["species_name", "variety", "analysis_group"], dropna=False)
        .agg(
            n_accessions=("seq_id", "size"),
            n_unique_accession_names=("accession_name", "nunique"),
            n_with_latitude=("has_latitude", "sum"),
            n_with_longitude=("has_longitude", "sum"),
            n_with_altitude=("has_altitude", "sum"),
        )
        .reset_index()
        .sort_values(["n_accessions", "species_name", "variety"], ascending=[False, True, True])
    )
    return group_df


def format_workbook(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="BFBFBF"))

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        max_widths = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                else:
                    cell.alignment = Alignment(vertical="top")

                if cell.value is None:
                    continue
                value_length = min(60, len(str(cell.value)) + 2)
                max_widths[cell.column] = max(max_widths.get(cell.column, 0), value_length)

        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        for column_idx, width in max_widths.items():
            ws.column_dimensions[get_column_letter(column_idx)].width = max(12, min(width, 60))

    wb.save(workbook_path)


def write_excel_workbook(workbook_path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            if df is None:
                pd.DataFrame().to_excel(writer, sheet_name=safe_name, index=False)
            else:
                df.to_excel(writer, sheet_name=safe_name, index=False)
    format_workbook(workbook_path)


def main() -> None:
    args = parse_args()

    project_dir = Path(args.project_dir).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve() if args.output_dir else (project_dir / "analysis_outputs")
    output_dir.mkdir(parents=True, exist_ok=True)

    file_map = locate_inputs(project_dir)
    input_inventory_df = build_input_inventory(file_map)

    print("Loading metadata")
    raw_meta_df, clean_meta_df = load_and_clean_metadata(file_map["metadata"])

    print("Inventorying syntenic alignments")
    synteny_df = inventory_synteny_tar(file_map["synteny"])

    sgc_result = parse_vcf_stream(
        file_map["sgC_vcf"],
        clean_meta_df,
        max_analysis_variants=args.max_analysis_variants_per_vcf,
        min_site_call_rate=args.min_site_call_rate,
        maf_threshold=args.maf_threshold,
        progress_every=args.progress_every,
        random_seed=args.random_seed,
        min_group_size=args.min_group_size,
    )

    sge_result = parse_vcf_stream(
        file_map["sgE_vcf"],
        clean_meta_df,
        max_analysis_variants=args.max_analysis_variants_per_vcf,
        min_site_call_rate=args.min_site_call_rate,
        maf_threshold=args.maf_threshold,
        progress_every=args.progress_every,
        random_seed=args.random_seed + 1,
        min_group_size=args.min_group_size,
    )

    dataset_summary_df = build_dataset_summary(clean_meta_df, sgc_result, sge_result, synteny_df)
    group_summary_df = build_group_summary(clean_meta_df)
    sample_overlap_df = build_sample_overlap(sgc_result, sge_result)
    subgenome_asymmetry_df = build_subgenome_asymmetry(clean_meta_df, sgc_result, sge_result)

    matched_sgc = set(sgc_result["sample_names"])
    matched_sge = set(sge_result["sample_names"])
    supp_s1_df = clean_meta_df.copy()
    supp_s1_df["present_in_sgC"] = supp_s1_df["seq_id"].isin(matched_sgc)
    supp_s1_df["present_in_sgE"] = supp_s1_df["seq_id"].isin(matched_sge)
    supp_s1_df["present_in_both"] = supp_s1_df["present_in_sgC"] & supp_s1_df["present_in_sgE"]
    supp_s1_df = supp_s1_df.merge(
        sgc_result["sample_qc_df"].add_prefix("sgC_").rename(columns={"sgC_sample_id": "seq_id"}),
        how="left",
        on="seq_id",
    )
    supp_s1_df = supp_s1_df.merge(
        sge_result["sample_qc_df"].add_prefix("sgE_").rename(columns={"sgE_sample_id": "seq_id"}),
        how="left",
        on="seq_id",
    )

    run_info_df = pd.DataFrame(
        {
            "parameter": [
                "project_dir",
                "output_dir",
                "max_analysis_variants_per_vcf",
                "min_site_call_rate",
                "maf_threshold",
                "min_group_size",
                "progress_every",
                "random_seed",
            ],
            "value": [
                str(project_dir),
                str(output_dir),
                args.max_analysis_variants_per_vcf,
                args.min_site_call_rate,
                args.maf_threshold,
                args.min_group_size,
                args.progress_every,
                args.random_seed,
            ],
        }
    )

    manuscript_workbook = output_dir / "coffee_introgression_manuscript_analysis.xlsx"
    supplementary_workbook = output_dir / "coffee_introgression_supplementary_data_s1.xlsx"

    manuscript_sheets = {
        "01_run_info": run_info_df,
        "02_input_inventory": input_inventory_df,
        "03_dataset_summary": dataset_summary_df,
        "04_group_summary": group_summary_df,
        "05_sgC_header": sgc_result["header_summary_df"],
        "06_sgE_header": sge_result["header_summary_df"],
        "07_sgC_contig_header": sgc_result["header_contig_df"],
        "08_sgE_contig_header": sge_result["header_contig_df"],
        "09_sgC_variant_summary": sgc_result["variant_summary_df"],
        "10_sgE_variant_summary": sge_result["variant_summary_df"],
        "11_sgC_filter_counts": sgc_result["filter_df"],
        "12_sgE_filter_counts": sge_result["filter_df"],
        "13_sgC_sample_match": sgc_result["sample_match_df"],
        "14_sgE_sample_match": sge_result["sample_match_df"],
        "15_sample_overlap": sample_overlap_df,
        "16_sgC_sample_qc": sgc_result["sample_qc_df"],
        "17_sgE_sample_qc": sge_result["sample_qc_df"],
        "18_sgC_group_qc": sgc_result["group_summary_df"],
        "19_sgE_group_qc": sge_result["group_summary_df"],
        "20_subgenome_delta": subgenome_asymmetry_df,
        "21_sgC_contig_qc": sgc_result["contig_df"],
        "22_sgE_contig_qc": sge_result["contig_df"],
        "23_sgC_pca_scores": sgc_result["pca_scores_df"],
        "24_sgE_pca_scores": sge_result["pca_scores_df"],
        "25_sgC_pca_var": sgc_result["pca_variance_df"],
        "26_sgE_pca_var": sge_result["pca_variance_df"],
        "27_sgC_ibs": sgc_result["ibs_df"].reset_index(),
        "28_sgE_ibs": sge_result["ibs_df"].reset_index(),
        "29_sgC_analysis_vars": sgc_result["analysis_variant_df"],
        "30_sgE_analysis_vars": sge_result["analysis_variant_df"],
        "31_sgC_top_contrasts": sgc_result["contrast_df"],
        "32_sgE_top_contrasts": sge_result["contrast_df"],
        "33_synteny_inventory": synteny_df,
    }

    supplementary_sheets = {
        "S1_accession_master": supp_s1_df,
        "S1_column_guide": pd.DataFrame(
            {
                "column_name": [
                    "seq_id",
                    "accession_name",
                    "species_name",
                    "variety",
                    "analysis_group",
                    "present_in_sgC",
                    "present_in_sgE",
                    "present_in_both",
                    "sgC_call_rate_biallelic_snps",
                    "sgE_call_rate_biallelic_snps",
                    "sgC_heterozygosity_rate_biallelic_snps",
                    "sgE_heterozygosity_rate_biallelic_snps",
                    "sgC_non_reference_rate_biallelic_snps",
                    "sgE_non_reference_rate_biallelic_snps",
                ],
                "description": [
                    "Sequencing identifier used to match metadata with VCF samples.",
                    "Accession or cultivar name.",
                    "Species name from the metadata workbook.",
                    "Normalized cultivation-status label.",
                    "Analysis-ready grouping label used in the exploratory contrasts.",
                    "Whether the accession is present in the sgC VCF.",
                    "Whether the accession is present in the sgE VCF.",
                    "Whether the accession is present in both subgenome VCF files.",
                    "Per-sample call rate across biallelic SNP sites in sgC.",
                    "Per-sample call rate across biallelic SNP sites in sgE.",
                    "Per-sample heterozygosity rate across called biallelic SNPs in sgC.",
                    "Per-sample heterozygosity rate across called biallelic SNPs in sgE.",
                    "Per-sample non-reference genotype rate across called biallelic SNPs in sgC.",
                    "Per-sample non-reference genotype rate across called biallelic SNPs in sgE.",
                ],
            }
        ),
    }

    print(f"Writing workbook: {manuscript_workbook}")
    write_excel_workbook(manuscript_workbook, manuscript_sheets)

    print(f"Writing workbook: {supplementary_workbook}")
    write_excel_workbook(supplementary_workbook, supplementary_sheets)

    manifest = {
        "project_dir": str(project_dir),
        "output_dir": str(output_dir),
        "manuscript_workbook": str(manuscript_workbook),
        "supplementary_workbook": str(supplementary_workbook),
        "input_files": {k: str(v) for k, v in file_map.items()},
    }

    manifest_path = output_dir / "coffee_introgression_output_manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as handle:
        json.dump(manifest, handle, indent=2)

    print("Analysis finished successfully.")
    print(f"Manifest written to: {manifest_path}")


if __name__ == "__main__":
    main()
