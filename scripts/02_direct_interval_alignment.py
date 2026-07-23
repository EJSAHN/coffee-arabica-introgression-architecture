#!/usr/bin/env python3
"""Direct sequence-alignment validation with tabular outputs only.

The script aligns submitted chromosome 4 source intervals to the expected
ET-39 pseudomolecules and writes workbooks, JSON, FASTA chunks, and BLAST TSV
files. It does not render manuscript figures.
"""
from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import math
import os
import platform
import re
import shutil
import stat
import subprocess
import sys
import tarfile
import urllib.request
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_VERSION = "1.0"
SOURCE_ACCESSION = "GCA_036785775.1"
NCBI_DOWNLOAD_URL = (
    "https://api.ncbi.nlm.nih.gov/datasets/v2/genome/accession/"
    + SOURCE_ACCESSION
    + "/download?include_annotation_type=GENOME_FASTA"
    + "&include_annotation_type=SEQUENCE_REPORT&hydrated=FULLY_HYDRATED"
)
BLAST_INDEX_URL = "https://ftp.ncbi.nlm.nih.gov/blast/executables/blast+/LATEST/"


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--project-dir", type=Path, required=True)
    p.add_argument("--output-dir", type=Path, required=True)
    p.add_argument("--intervals-csv", type=Path, required=False)
    p.add_argument("--threads", type=int, default=4)
    p.add_argument("--chunk-size", type=int, default=200_000)
    p.add_argument("--chunk-overlap", type=int, default=20_000)
    p.add_argument("--min-identity", type=float, default=85.0)
    p.add_argument("--min-hsp-length", type=int, default=1_000)
    p.add_argument("--force-download", action="store_true")
    p.add_argument("--reuse-existing-blast", action="store_true", help="Reuse non-empty blast_sgC.tsv/blast_sgE.tsv already present in the output directory")
    p.add_argument("--blastn-path", type=Path, help="Optional explicit path to blastn/blastn.exe")
    p.add_argument("--makeblastdb-path", type=Path, help="Optional explicit path to makeblastdb/makeblastdb.exe")
    return p.parse_args()


def normalize_seqid(value: str) -> str:
    text = "" if value is None else str(value)
    match = re.search(r"(NC_\d+\.\d+)", text)
    if match:
        return match.group(1)
    return text.strip().strip("|")


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_seqid(s).lower())


def find_file(root: Path, names: list[str]) -> Optional[Path]:
    for name in names:
        p = root / name
        if p.exists():
            return p
    for p in root.rglob("*"):
        if p.is_file() and p.name in names:
            return p
    return None


def read_vcf_contig_lengths(path: Path) -> dict[str, int]:
    out = {}
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as h:
        for line in h:
            if not line.startswith("##"):
                break
            if line.startswith("##contig=<"):
                m_id = re.search(r"ID=([^,>]+)", line)
                m_len = re.search(r"length=(\d+)", line)
                if m_id and m_len:
                    out[m_id.group(1)] = int(m_len.group(1))
    return out


def download(url: str, dest: Path):
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(dest.suffix + ".part")
    req = urllib.request.Request(url, headers={"User-Agent": "coffee-introgression-alignment/1.0"})
    with urllib.request.urlopen(req, timeout=180) as r, tmp.open("wb") as w:
        total = int(r.headers.get("Content-Length", 0) or 0)
        got = 0
        while True:
            block = r.read(1024 * 1024)
            if not block:
                break
            w.write(block)
            got += len(block)
            if total and got % (100 * 1024 * 1024) < len(block):
                print(f"  downloaded {got/1e6:.1f}/{total/1e6:.1f} MB", flush=True)
    tmp.replace(dest)


def ensure_source_package(resources: Path, force: bool = False) -> tuple[Path, Optional[Path]]:
    pkg = resources / f"{SOURCE_ACCESSION}_dataset.zip"
    extract = resources / SOURCE_ACCESSION
    if force and pkg.exists():
        pkg.unlink()
    if not pkg.exists():
        print(f"Downloading source assembly {SOURCE_ACCESSION} from NCBI Datasets...")
        download(NCBI_DOWNLOAD_URL, pkg)
    if force and extract.exists():
        shutil.rmtree(extract)
    if not extract.exists():
        extract.mkdir(parents=True)
        with zipfile.ZipFile(pkg) as z:
            z.extractall(extract)
    fastas = [p for p in extract.rglob("*") if p.is_file() and re.search(r"\.(fna|fa|fasta)(\.gz)?$", p.name, re.I)]
    if not fastas:
        raise RuntimeError(f"No genome FASTA found in {pkg}")
    fasta = max(fastas, key=lambda p: p.stat().st_size)
    reports = [p for p in extract.rglob("sequence_report.jsonl") if p.is_file()]
    return fasta, (reports[0] if reports else None)


def latest_windows_blast_asset() -> str:
    req = urllib.request.Request(BLAST_INDEX_URL, headers={"User-Agent": "coffee-introgression-alignment/1.0"})
    html = urllib.request.urlopen(req, timeout=60).read().decode("utf-8", "replace")
    assets = re.findall(r'href="([^"]*ncbi-blast-([0-9.]+)\+-x64-win64\.tar\.gz)"', html)
    if not assets:
        raise RuntimeError("Could not identify the Windows BLAST+ archive from the NCBI LATEST directory")

    def key(item):
        return tuple(int(value) for value in item[1].split("."))

    return sorted(assets, key=key)[-1][0].split("/")[-1]


def _validated_blast_pair(blastn: Path | None, makeblastdb: Path | None) -> tuple[Path, Path] | None:
    if blastn is None or makeblastdb is None:
        return None
    blastn = blastn.expanduser().resolve()
    makeblastdb = makeblastdb.expanduser().resolve()
    if blastn.is_file() and makeblastdb.is_file():
        return blastn, makeblastdb
    return None


def ensure_blast(
    resources: Path,
    force: bool = False,
    blastn_path: Path | None = None,
    makeblastdb_path: Path | None = None,
) -> tuple[Path, Path]:
    # 1. Explicit paths supplied by the user.
    explicit = _validated_blast_pair(blastn_path, makeblastdb_path)
    if explicit is not None:
        return explicit
    if (blastn_path is None) != (makeblastdb_path is None):
        raise RuntimeError("Provide both --blastn-path and --makeblastdb-path, or neither.")

    # 2. Existing installation on PATH (portable Linux/macOS/Windows route).
    blastn_cmd = shutil.which("blastn")
    makeblastdb_cmd = shutil.which("makeblastdb")
    path_pair = _validated_blast_pair(Path(blastn_cmd) if blastn_cmd else None, Path(makeblastdb_cmd) if makeblastdb_cmd else None)
    if path_pair is not None:
        return path_pair

    # 3. Previously cached executables in the project resources directory.
    blast_names = {"blastn", "blastn.exe"}
    db_names = {"makeblastdb", "makeblastdb.exe"}
    for blastn in resources.rglob("*"):
        if blastn.is_file() and blastn.name in blast_names:
            makeblastdb = blastn.parent / ("makeblastdb.exe" if blastn.suffix.lower() == ".exe" else "makeblastdb")
            if makeblastdb.name in db_names and makeblastdb.is_file():
                return blastn.resolve(), makeblastdb.resolve()

    # 4. Automated download is intentionally limited to Windows because the
    #    official archive naming and binary layout are platform specific.
    if platform.system().lower() != "windows":
        raise RuntimeError(
            "NCBI BLAST+ was not found on PATH. Install blastn and makeblastdb "
            "with your system package manager or Conda/Bioconda, then rerun. "
            "Alternatively provide --blastn-path and --makeblastdb-path explicitly."
        )

    asset = latest_windows_blast_asset()
    archive = resources / asset
    if force and archive.exists():
        archive.unlink()
    if not archive.exists():
        print(f"Downloading official NCBI BLAST+ for Windows: {asset}")
        download(BLAST_INDEX_URL + asset, archive)
    out = resources / "blast_plus"
    if force and out.exists():
        shutil.rmtree(out)
    out.mkdir(parents=True, exist_ok=True)
    with tarfile.open(archive, "r:gz") as tf:
        tf.extractall(out)
    blastn = next((candidate for candidate in out.rglob("blastn.exe") if candidate.is_file()), None)
    if blastn is None:
        raise RuntimeError("blastn.exe was not found after extracting the Windows BLAST+ archive")
    makeblastdb = blastn.parent / "makeblastdb.exe"
    if not makeblastdb.exists():
        raise RuntimeError("makeblastdb.exe was not found after extracting the Windows BLAST+ archive")
    return blastn.resolve(), makeblastdb.resolve()

def fasta_iter(path: Path):
    opener = gzip.open if path.name.lower().endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as h:
        name = None
        desc = None
        seq = []
        for line in h:
            if line.startswith(">"):
                if name is not None:
                    yield name, desc, "".join(seq)
                desc = line[1:].strip()
                name = desc.split()[0]
                seq = []
            else:
                seq.append(line.strip())
        if name is not None:
            yield name, desc, "".join(seq)


def choose_source_record(source_fasta: Path, source_alias: str, expected_len: int) -> tuple[str, str, str, str]:
    alias_n = norm(source_alias)
    scaffold_tokens = re.findall(r"(?:scaffold|hrscaf)_?\d+", source_alias, re.I)
    candidates = []
    for name, desc, seq in fasta_iter(source_fasta):
        score = 0
        d = norm(desc)
        if alias_n and alias_n in d:
            score += 100
        for token in scaffold_tokens:
            if norm(token) in d:
                score += 20
        diff = abs(len(seq) - expected_len)
        if diff == 0:
            score += 80
        elif expected_len and diff / expected_len < 0.001:
            score += 30
        candidates.append((score, diff, name, desc, seq))
    candidates.sort(key=lambda x: (-x[0], x[1]))
    if not candidates:
        raise RuntimeError("Source assembly FASTA contains no records")
    best = candidates[0]
    exact_len = [x for x in candidates if x[1] == 0]
    if best[0] < 50 and len(exact_len) != 1:
        top = [(x[2], len(x[4]), x[0], x[1]) for x in candidates[:10]]
        raise RuntimeError(f"Could not uniquely map VCF contig {source_alias} length {expected_len}; top candidates: {top}")
    method = "alias_and_length" if best[0] >= 100 else ("exact_length" if best[1] == 0 else "best_available")
    return best[2], best[3], best[4], method


def write_fasta(path: Path, records: Iterable[tuple[str, str]]):
    with path.open("w", encoding="utf-8") as w:
        for name, seq in records:
            w.write(f">{name}\n")
            for i in range(0, len(seq), 80):
                w.write(seq[i:i+80] + "\n")


def make_chunks(seq: str, source_start: int, source_end: int, chunk_size: int, overlap: int):
    interval = seq[source_start-1:source_end]
    step = max(1, chunk_size - overlap)
    rows = []
    records = []
    idx = 0
    for local0 in range(0, len(interval), step):
        local1 = min(len(interval), local0 + chunk_size)
        if local1 - local0 < 1000:
            continue
        idx += 1
        abs_start = source_start + local0
        abs_end = source_start + local1 - 1
        name = f"chunk{idx:04d}|src={abs_start}-{abs_end}"
        records.append((name, interval[local0:local1]))
        rows.append({"query": name, "chunk_index": idx, "source_start": abs_start, "source_end": abs_end, "chunk_length": local1-local0})
        if local1 == len(interval):
            break
    return records, pd.DataFrame(rows)


def run(
    cmd: list[str],
    stdout: Optional[Path] = None,
    stderr: Optional[Path] = None,
    cwd: Optional[Path] = None,
):
    print("RUN:", " ".join(f'"{x}"' if " " in x else x for x in cmd), flush=True)
    if cwd is not None:
        print("CWD:", str(cwd), flush=True)
    out_handle = stdout.open("w", encoding="utf-8") if stdout else subprocess.DEVNULL
    err_handle = stderr.open("w", encoding="utf-8") if stderr else subprocess.PIPE
    try:
        res = subprocess.run(
            cmd,
            stdout=out_handle,
            stderr=err_handle,
            text=True,
            check=False,
            cwd=str(cwd) if cwd else None,
        )
    finally:
        if stdout and hasattr(out_handle, "close"):
            out_handle.close()
        if stderr and hasattr(err_handle, "close"):
            err_handle.close()
    if res.returncode != 0:
        msg = ""
        if stderr and stderr.exists():
            try:
                msg = stderr.read_text(encoding="utf-8", errors="replace")
            except OSError:
                msg = ""
        elif isinstance(res.stderr, str):
            msg = res.stderr
        raise RuntimeError(
            f"Command failed ({res.returncode}): {' '.join(cmd)}\n"
            f"Working directory: {cwd or Path.cwd()}\n{msg}"
        )


def clear_windows_readonly(path: Path) -> None:
    """Best-effort removal of Windows read-only attributes."""
    if os.name != "nt":
        return
    try:
        subprocess.run(
            ["attrib", "-R", str(path)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
    except OSError:
        pass


def verify_writable_directory(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    clear_windows_readonly(path)
    try:
        path.chmod(stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
    except OSError:
        pass
    probe = path / "_python_write_probe.tmp"
    try:
        probe.write_text("write test\n", encoding="utf-8")
    except OSError as exc:
        raise RuntimeError(f"Python cannot write to BLAST database directory {path}: {exc}") from exc
    finally:
        try:
            probe.unlink()
        except OSError:
            pass


def locate_or_build_blast_db(
    makeblastdb: Path,
    target_fasta: Path,
    subgenome: str,
    resources: Path,
    output_dir: Path,
) -> tuple[Path, Path]:
    """Build the BLAST database in a verified writable directory.

    NCBI BLAST on Windows checks the output directory independently.  To avoid
    false permission failures caused by inherited attributes, makeblastdb is
    run *inside* the database directory with a relative database prefix.  If
    the project cache is rejected, the function automatically falls back to
    LOCALAPPDATA and then TEMP.
    """
    candidates: list[Path] = [
        resources / "blastdb_cache_v41" / subgenome,
    ]
    local_appdata = os.environ.get("LOCALAPPDATA")
    temp_dir = os.environ.get("TEMP") or os.environ.get("TMP")
    if local_appdata:
        candidates.append(Path(local_appdata) / "coffee_revision_blastdb_v41" / subgenome)
    if temp_dir:
        candidates.append(Path(temp_dir) / "coffee_revision_blastdb_v41" / subgenome)

    errors: list[str] = []
    for attempt, dbdir in enumerate(candidates, start=1):
        try:
            verify_writable_directory(dbdir)
        except Exception as exc:
            errors.append(f"{dbdir}: directory preparation failed: {exc}")
            continue

        prefix_name = f"ET39_{subgenome}"
        dbprefix = dbdir / prefix_name
        if dbprefix.with_suffix(".ndb").exists() or dbprefix.with_suffix(".nin").exists():
            print(f"Using existing BLAST database: {dbprefix}", flush=True)
            return dbdir, dbprefix

        log = output_dir / f"makeblastdb_{subgenome}_attempt{attempt}.stderr.log"
        # Remove stale partial database files before rebuilding.
        for partial in dbdir.glob(prefix_name + "*"):
            try:
                partial.unlink()
            except OSError:
                pass

        try:
            run(
                [
                    str(makeblastdb),
                    "-in", str(target_fasta),
                    "-dbtype", "nucl",
                    "-parse_seqids",
                    "-out", prefix_name,
                ],
                stderr=log,
                cwd=dbdir,
            )
        except Exception as exc:
            errors.append(f"{dbdir}: {exc}")
            continue

        if dbprefix.with_suffix(".ndb").exists() or dbprefix.with_suffix(".nin").exists():
            print(f"Created BLAST database: {dbprefix}", flush=True)
            return dbdir, dbprefix
        errors.append(f"{dbdir}: makeblastdb returned success but no .ndb/.nin file was found")

    joined = "\n\n".join(errors)
    raise RuntimeError(
        f"Could not create a writable BLAST database for {subgenome}. "
        f"Tried {len(candidates)} locations.\n{joined}"
    )


def merge_intervals(intervals):
    vals = sorted((min(int(a),int(b)), max(int(a),int(b))) for a,b in intervals)
    if not vals: return []
    out=[vals[0]]
    for a,b in vals[1:]:
        x,y=out[-1]
        if a <= y+1: out[-1]=(x,max(y,b))
        else: out.append((a,b))
    return out


def total_len(intervals):
    return sum(b-a+1 for a,b in merge_intervals(intervals))


def unaligned_segments(start, end, covered):
    cov=merge_intervals(covered)
    out=[]; cursor=start
    for a,b in cov:
        a=max(a,start); b=min(b,end)
        if a>cursor: out.append((cursor,a-1))
        cursor=max(cursor,b+1)
    if cursor<=end: out.append((cursor,end))
    return out


def parse_blast(path: Path, chunks: pd.DataFrame, min_identity: float, min_hsp: int):
    cols = ["query","qlen","qstart","qend","target","slen","sstart","send","pident","length","bitscore","evalue"]
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame(columns=cols)
    df=pd.read_csv(path, sep="\t", names=cols)
    for c in ["qlen","qstart","qend","slen","sstart","send","length"]: df[c]=pd.to_numeric(df[c],errors="coerce")
    for c in ["pident","bitscore","evalue"]: df[c]=pd.to_numeric(df[c],errors="coerce")
    df=df[(df["pident"]>=min_identity)&(df["length"]>=min_hsp)].copy()
    meta=chunks.set_index("query")
    df["chunk_source_start"]=df["query"].map(meta["source_start"])
    df["source_hsp_start"]=df["chunk_source_start"]+df[["qstart","qend"]].min(axis=1)-1
    df["source_hsp_end"]=df["chunk_source_start"]+df[["qstart","qend"]].max(axis=1)-1
    df["target_hsp_start"]=df[["sstart","send"]].min(axis=1)
    df["target_hsp_end"]=df[["sstart","send"]].max(axis=1)
    df["orientation"]=np.where((df["qend"]-df["qstart"])*(df["send"]-df["sstart"])>=0,"+","-")
    return df


def summarize(df: pd.DataFrame, source_start: int, source_end: int, expected_target: str):
    span=source_end-source_start+1
    if df.empty:
        summary = {
            "status": "FAIL",
            "source_interval_start": source_start,
            "source_interval_end": source_end,
            "source_interval_span_bp": span,
            "dominant_target_contig": None,
            "expected_target_contig": expected_target,
            "expected_target_match": False,
            "aligned_source_bp": 0,
            "aligned_source_fraction": 0.0,
            "all_target_union_coverage_fraction": 0.0,
            "largest_alternative_target_fraction": 0.0,
            "mapped_target_start": None,
            "mapped_target_end": None,
            "mapped_target_envelope_span_bp": None,
            "weighted_identity": None,
            "orientation_set": None,
            "n_dominant_hsps": 0,
            "n_dominant_chunks": 0,
            "n_target_contigs": 0,
        }
        return summary, pd.DataFrame(), pd.DataFrame([
            {"unaligned_start": source_start, "unaligned_end": source_end, "unaligned_span_bp": span}
        ])
    # Best HSP per query-target pair to limit repetitive double counting.
    best=df.sort_values(["bitscore","length"],ascending=False).drop_duplicates(["query","target"])
    target_rows=[]
    for target,g in best.groupby("target"):
        cov=total_len(list(zip(g["source_hsp_start"],g["source_hsp_end"])))
        target_rows.append({"target":target,"covered_source_bp":cov,"coverage_fraction":cov/span,"median_identity":float(g["pident"].median()),"n_chunks":int(g["query"].nunique()),"n_hsps":len(g)})
    targets=pd.DataFrame(target_rows).sort_values(["covered_source_bp","median_identity"],ascending=False)
    dominant=str(targets.iloc[0]["target"])
    dom=best[best["target"]==dominant].copy()
    source_cov=total_len(list(zip(dom["source_hsp_start"],dom["source_hsp_end"])))
    all_cov=total_len(list(zip(best["source_hsp_start"],best["source_hsp_end"])))
    alt_cov=0
    for target,g in best[best["target"]!=dominant].groupby("target"):
        alt_cov=max(alt_cov,total_len(list(zip(g["source_hsp_start"],g["source_hsp_end"]))))
    expected_match = norm(dominant)==norm(expected_target)
    # target range from dominant HSPs; report block envelope only, not implied gap alignment.
    target_start=int(dom["target_hsp_start"].min()); target_end=int(dom["target_hsp_end"].max())
    orient=";".join(sorted(set(dom["orientation"])))
    median_identity=float(np.average(dom["pident"],weights=dom["length"]))
    cov_frac=source_cov/span
    alt_frac=min(1.0,alt_cov/span)
    if expected_match and cov_frac>=0.70 and alt_frac<=0.10 and median_identity>=90:
        status="PASS"
    elif expected_match and cov_frac>=0.40 and median_identity>=85:
        status="WARN"
    else:
        status="FAIL"
    summary={
        "status":status,"source_interval_start":source_start,"source_interval_end":source_end,"source_interval_span_bp":span,
        "dominant_target_contig":dominant,"expected_target_contig":expected_target,"expected_target_match":expected_match,
        "aligned_source_bp":source_cov,"aligned_source_fraction":cov_frac,"all_target_union_coverage_fraction":all_cov/span,
        "largest_alternative_target_fraction":alt_frac,"mapped_target_start":target_start,"mapped_target_end":target_end,
        "mapped_target_envelope_span_bp":target_end-target_start+1,"weighted_identity":median_identity,
        "orientation_set":orient,"n_dominant_hsps":len(dom),"n_dominant_chunks":int(dom["query"].nunique()),"n_target_contigs":int(targets.shape[0]),
    }
    gaps=pd.DataFrame([{"unaligned_start":a,"unaligned_end":b,"unaligned_span_bp":b-a+1} for a,b in unaligned_segments(source_start,source_end,list(zip(dom["source_hsp_start"],dom["source_hsp_end"])))])
    return summary,targets,gaps


def write_excel(path: Path, sheets: dict[str,pd.DataFrame]):
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name,df in sheets.items():
            if df is None or df.empty:
                pd.DataFrame({"message":["No records"]}).to_excel(xw,sheet_name=name[:31],index=False)
            else:
                df.to_excel(xw,sheet_name=name[:31],index=False)
        for ws in xw.book.worksheets:
            ws.freeze_panes="A2"
            for cell in ws[1]:
                cell.font=Font(bold=True)
                cell.fill=PatternFill("solid",fgColor="D9EAF7")
            for col in range(1,ws.max_column+1):
                vals=[str(ws.cell(r,col).value or "") for r in range(1,min(ws.max_row,100)+1)]
                ws.column_dimensions[get_column_letter(col)].width=min(60,max(10,max(map(len,vals))+2))


def main():
    a=parse_args(); root=a.project_dir.resolve(); out=a.output_dir.resolve(); out.mkdir(parents=True,exist_ok=True)
    resources=root/"revision_alignment_resources"; resources.mkdir(exist_ok=True)
    if a.intervals_csv is not None:
        intervals = pd.read_csv(a.intervals_csv)
    else:
        intervals = pd.DataFrame([
            {
                "subgenome": "sgC",
                "source_contig": "chr_D_sg_C_(Scaffold_1;HRSCAF_2)",
                "source_start": 6804956,
                "source_end": 30148002,
                "expected_target_seqid": "NC_092316.1",
                "expected_target_label": "4c",
            },
            {
                "subgenome": "sgE",
                "source_contig": "chr_D_sg_E_(Scaffold_2;HRSCAF_3)",
                "source_start": 7028545,
                "source_end": 30145533,
                "expected_target_seqid": "NC_092317.1",
                "expected_target_label": "4e",
            },
        ])
    sgc_vcf=find_file(root,["Arabica_sgC.TIP.BB.vcf.gz"]); sge_vcf=find_file(root,["Arabica_sgE.TIP.BB.vcf.gz"])
    target_sgc=find_file(root,["C_arabica_ET39_sgC.fasta"]); target_sge=find_file(root,["C_arabica_ET39_sgE.fasta"])
    if not all([sgc_vcf,sge_vcf,target_sgc,target_sge]):
        raise RuntimeError("Required VCF or target FASTA missing")
    source_fasta,seq_report=ensure_source_package(resources,a.force_download)
    blastn,makeblastdb=ensure_blast(resources,a.force_download,a.blastn_path,a.makeblastdb_path)
    inputs=[]; source_map=[]; all_hsps=[]; all_targets=[]; all_gaps=[]; summaries=[]; chunk_tables=[]
    for row in intervals.to_dict("records"):
        sub=row["subgenome"]; alias=row["source_contig"]; s0=int(row["source_start"]); s1=int(row["source_end"]); expected=row["expected_target_seqid"]
        vcf=sgc_vcf if sub=="sgC" else sge_vcf; target=target_sgc if sub=="sgC" else target_sge
        lengths=read_vcf_contig_lengths(vcf); expected_len=lengths.get(alias)
        if expected_len is None: raise RuntimeError(f"VCF contig not found: {alias}")
        src_name,src_desc,src_seq,method=choose_source_record(source_fasta,alias,expected_len)
        if len(src_seq)<s1: raise RuntimeError(f"Source sequence {src_name} length {len(src_seq)} shorter than interval {s1}")
        source_map.append({"subgenome":sub,"vcf_contig":alias,"vcf_contig_length":expected_len,"source_fasta_record":src_name,"source_fasta_description":src_desc,"source_fasta_length":len(src_seq),"mapping_method":method})
        records,chunks=make_chunks(src_seq,s0,s1,a.chunk_size,a.chunk_overlap); chunks.insert(0,"subgenome",sub); chunk_tables.append(chunks)
        qfa=out/f"{sub}_submitted_interval_chunks.fasta"; write_fasta(qfa,records)
        dbdir,dbprefix=locate_or_build_blast_db(
            makeblastdb=makeblastdb,
            target_fasta=target,
            subgenome=sub,
            resources=resources,
            output_dir=out,
        )
        tab=out/f"blast_{sub}.tsv"
        if a.reuse_existing_blast and tab.exists() and tab.stat().st_size > 0:
            print(f"Reusing existing BLAST output: {tab}", flush=True)
        else:
            run(
                [str(blastn),"-task","megablast","-query",str(qfa),"-db",dbprefix.name,"-out",str(tab),"-outfmt","6 qseqid qlen qstart qend sseqid slen sstart send pident length bitscore evalue","-num_threads",str(a.threads),"-max_target_seqs","20","-max_hsps","20","-evalue","1e-20","-dust","yes"],
                stderr=out/f"blast_{sub}.stderr.log",
                cwd=dbdir,
            )
        print(f"Parsing BLAST output for {sub}: {tab}", flush=True)
        hsps=parse_blast(tab,chunks,a.min_identity,a.min_hsp_length)
        hsps.insert(0,"subgenome",sub)
        all_hsps.append(hsps)
        print(f"Retained {len(hsps):,} HSPs for {sub}; summarizing...", flush=True)
        summary,targets,gaps=summarize(hsps,s0,s1,expected)
        summary.update({"subgenome":sub,"source_contig":alias,"source_fasta_record":src_name,"alignment_method":"NCBI_BLAST_megablast_chunks"})
        summaries.append(summary)
        print(json.dumps({"subgenome": sub, **summary}, indent=2), flush=True)
        if not targets.empty:
            targets.insert(0,"subgenome",sub)
            all_targets.append(targets)
        if not gaps.empty:
            gaps.insert(0,"subgenome",sub)
            all_gaps.append(gaps)
        inputs.append({"resource":f"source_assembly_{SOURCE_ACCESSION}","path":str(source_fasta),"sequence_report":str(seq_report) if seq_report else None})
        inputs.append({"resource":f"target_{sub}_fasta","path":str(target),"sequence_report":None})
    summary_df=pd.DataFrame(summaries)
    overall="PASS" if len(summary_df) and (summary_df.status=="PASS").all() else ("WARN" if len(summary_df) and (summary_df.status.isin(["PASS","WARN"])).all() else "FAIL")
    sheets={"Validation_status":pd.DataFrame([{"overall_status":overall}]),"Input_inventory":pd.DataFrame(inputs),"Source_contig_mapping":pd.DataFrame(source_map),"Submitted_intervals":intervals,"Interval_summary":summary_df,"Target_summary":pd.concat(all_targets,ignore_index=True) if all_targets else pd.DataFrame(),"Chunk_manifest":pd.concat(chunk_tables,ignore_index=True),"Blast_HSPs":pd.concat(all_hsps,ignore_index=True) if all_hsps else pd.DataFrame(),"Unaligned_segments":pd.concat(all_gaps,ignore_index=True) if all_gaps else pd.DataFrame()}
    write_excel(out/"submitted_interval_alignment_validation.xlsx",sheets)
    status={"script_version":SCRIPT_VERSION,"overall_status":overall,"source_assembly":SOURCE_ACCESSION,"subgenomes":summaries}
    (out/"submitted_interval_alignment_status.json").write_text(json.dumps(status,indent=2),encoding="utf-8")
    print(json.dumps(status,indent=2))
    print("Workbook:",out/"submitted_interval_alignment_validation.xlsx")

if __name__=="__main__":
    main()
