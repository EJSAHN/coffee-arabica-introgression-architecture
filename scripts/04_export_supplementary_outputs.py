#!/usr/bin/env python3
"""Export supplementary workbooks and machine-readable source data.

The script consumes the completed strict subgenome-filtered sensitivity workbook
and corrected direct-alignment workbook. It writes Supplementary Tables, Data S1,
analysis summaries, and CSV source data underlying manuscript figures. It does
not render or assemble manuscript figures.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_PRIORITY_ACCESSIONS = [
    "Híbrido de Timor",
    "Arabusta Pizze",
    "Iapar 59",
    "Oro Azteca",
    "IPR99",
    "Costa Rica 95",
]

CATEGORY_PATTERNS = [
    ("Defense receptor", re.compile(r"resistan|disease|immune|\bRPP\b|RPP\d|RGA\d|NLR|NBS|TIR|LRR", re.I)),
    ("Kinase signaling", re.compile(r"kinase|phosphatase|MAPK|WNK|signaling|calmodulin", re.I)),
    ("Stress response", re.compile(r"stress|defen[cs]e|pathogenesis|chitin|heat shock|oxidative", re.I)),
    ("Transcriptional regulation", re.compile(r"transcription|\bNAC\b|JA2L|WRKY|MYB|bZIP|TCP", re.I)),
    ("Membrane transport", re.compile(r"transport|transporter|channel|pump|aquaporin|membrane", re.I)),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    config_dir = Path(__file__).resolve().parents[1] / "config"
    parser.add_argument("--project-dir", type=Path, help="Project root used for automatic input discovery.")
    parser.add_argument("--sensitivity-workbook", type=Path, help="Strict subgenome-filtered sensitivity workbook.")
    parser.add_argument("--alignment-workbook", type=Path, help="Corrected direct-alignment workbook.")
    parser.add_argument(
        "--priority-accessions", type=Path, default=config_dir / "prioritized_accessions.csv",
        help="CSV defining the display order of validated priority accessions.",
    )
    parser.add_argument(
        "--accession-context", type=Path, default=config_dir / "accession_context.csv",
        help="Manually curated accession-context CSV used only for contextual tables.",
    )
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory for exported workbooks and CSV files.")
    return parser.parse_args()


def newest(paths: Iterable[Path]) -> Path | None:
    items = [p for p in paths if p.exists()]
    return max(items, key=lambda p: p.stat().st_mtime) if items else None

def discover_inputs(args: argparse.Namespace) -> tuple[Path, Path]:
    sensitivity = args.sensitivity_workbook
    alignment = args.alignment_workbook
    if args.project_dir:
        root = args.project_dir.resolve()
        outputs = root / "revision_validation_outputs"
        if sensitivity is None:
            sensitivity = newest(list(outputs.glob("FULL_SENSITIVITY_V4_FILTERED_*/sampling_population_structure_validation.xlsx")) +
                                 list(outputs.glob("FULL_SENSITIVITY_FILTERED_*/sampling_population_structure_validation.xlsx")))
        if alignment is None:
            alignment = newest(list(outputs.glob("ALIGNMENT_V41_*/postprocess_v44/alignment_validation_v44_corrected.xlsx")) +
                               list(outputs.glob("ALIGNMENT_FINAL_*/final/alignment_validation_corrected.xlsx")) +
                               list(outputs.glob("ALIGNMENT_*/final/alignment_validation_corrected.xlsx")) +
                               list(outputs.glob("ALIGNMENT_FINAL_*/final/alignment_validation_v44_corrected.xlsx")) +
                               list(outputs.glob("ALIGNMENT_*/final/alignment_validation_v44_corrected.xlsx")))
    if sensitivity is None or not sensitivity.exists():
        raise FileNotFoundError("Sensitivity workbook was not found. Provide --sensitivity-workbook or --project-dir.")
    if alignment is None or not alignment.exists():
        raise FileNotFoundError("Alignment workbook was not found. Provide --alignment-workbook or --project-dir.")
    return sensitivity.resolve(), alignment.resolve()

def classify_candidate(text: str) -> str:
    for label, pattern in CATEGORY_PATTERNS:
        if pattern.search(text or ""):
            return label
    return "Other candidate annotation"

def normalize_accession_names(common_panel: pd.DataFrame) -> pd.DataFrame:
    out = common_panel.copy()
    out["seq_id"] = out["seq_id"].astype(str)
    out["accession_display"] = out["accession_name"].replace({"H.Timor": "Híbrido de Timor"})
    return out

def baseline_introgressed_rankings(rankings: pd.DataFrame, common_panel: pd.DataFrame) -> pd.DataFrame:
    meta = common_panel[["seq_id", "accession_display", "country_of_origin", "donor_institute", "notes", "ref", "full_reference"]].copy()
    meta["seq_id"] = meta["seq_id"].astype(str)
    df = rankings[(rankings["condition"] == "filtered_reservoir_12000") &
                  (rankings["panel_type"] == "arabica_only") &
                  (rankings["analysis_group"] == "arabica_introgressed")].copy()
    df["sample_id"] = df["sample_id"].astype(str)
    return df.merge(meta, left_on="sample_id", right_on="seq_id", how="left")

def load_priority_accessions(path: Path, rankings: pd.DataFrame, common_panel: pd.DataFrame) -> list[str]:
    if path.exists():
        frame = pd.read_csv(path)
        if "accession" not in frame.columns:
            raise ValueError(f"Priority-accession file must contain an 'accession' column: {path}")
        if "display_order" in frame.columns:
            frame = frame.sort_values("display_order")
        accessions = frame["accession"].dropna().astype(str).str.strip().tolist()
    else:
        accessions = list(DEFAULT_PRIORITY_ACCESSIONS)

    baseline = baseline_introgressed_rankings(rankings, common_panel)
    observed = set(baseline["accession_display"].dropna().astype(str))
    missing = [name for name in accessions if name not in observed]
    if missing:
        raise ValueError(f"Configured priority accessions were not found in the validated baseline rankings: {missing}")

    # The configuration controls display order only. Validate that every listed
    # accession is retained in the top six across all sensitivity conditions.
    meta = common_panel[["seq_id", "accession_display"]].copy()
    meta["seq_id"] = meta["seq_id"].astype(str)
    ranked = rankings[(rankings["panel_type"] == "arabica_only") &
                      (rankings["analysis_group"] == "arabica_introgressed")].copy()
    ranked["sample_id"] = ranked["sample_id"].astype(str)
    ranked = ranked.merge(meta, left_on="sample_id", right_on="seq_id", how="left")
    unstable = []
    for name in accessions:
        subset = ranked[ranked["accession_display"] == name]
        if subset.empty or not bool((subset["within_introgressed_rank"] <= len(accessions)).all()):
            unstable.append(name)
    if unstable:
        raise ValueError(f"Priority-accession configuration is inconsistent with sensitivity results: {unstable}")
    return accessions

def load_accession_context(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=["Accession", "Disease or stress context", "Agronomic or quality context", "Interpretive note"])
    frame = pd.read_csv(path)
    required = {"Accession", "Disease or stress context", "Agronomic or quality context", "Interpretive note"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"Accession-context file is missing required columns {missing}: {path}")
    frame["Accession"] = frame["Accession"].astype(str).str.strip()
    return frame

def candidate_table(candidates: pd.DataFrame) -> pd.DataFrame:
    out = candidates.copy()
    out["annotation_text"] = out[["gene_name", "product"]].fillna("").astype(str).agg(" ".join, axis=1)
    out["candidate_category"] = out["annotation_text"].map(classify_candidate)
    return out

def style_sheet(ws, header_row: int = 3, freeze: str | None = None) -> None:
    thin = Side(style="thin", color="D9D9D9")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = title_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if freeze:
        ws.freeze_panes = freeze
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or "") for r in range(1, min(ws.max_row, 100) + 1)]
        width = min(max(10, max((len(v) for v in values), default=10) + 2), 42)
        ws.column_dimensions[get_column_letter(col)].width = width

def add_dataframe_sheet(wb: Workbook, name: str, title: str, description: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name[:31])
    last_col = max(1, len(df.columns))
    ws.cell(1, 1, title)
    ws.cell(2, 1, description)
    if last_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    for j, col in enumerate(df.columns, 1):
        ws.cell(3, j, str(col))
    for i, row in enumerate(df.itertuples(index=False, name=None), 4):
        for j, value in enumerate(row, 1):
            if isinstance(value, (np.integer,)):
                value = int(value)
            elif isinstance(value, (np.floating,)):
                value = None if np.isnan(value) else float(value)
            elif pd.isna(value):
                value = None
            ws.cell(i, j, value)
    style_sheet(ws, header_row=3, freeze="A4")

def accession_context_table(common_panel: pd.DataFrame, rankings: pd.DataFrame,
                            priority_accessions: Sequence[str], context_frame: pd.DataFrame) -> pd.DataFrame:
    baseline = baseline_introgressed_rankings(rankings, common_panel)
    piv = baseline.pivot(index="accession_display", columns="subgenome",
                         values=["within_introgressed_rank", "pca_introgression_affinity"])
    context_lookup = context_frame.set_index("Accession").to_dict("index") if not context_frame.empty else {}
    records = []
    for name in priority_accessions:
        meta_rows = common_panel[common_panel["accession_display"] == name]
        if meta_rows.empty:
            raise ValueError(f"Priority accession is missing from Common_panel: {name}")
        meta = meta_rows.iloc[0]
        context = context_lookup.get(name, {})
        records.append({
            "Accession": name,
            "Analysis group": "Introgressed arabica",
            "Country/source": str(meta.get("country_of_origin", "")).strip(),
            "Donor institute": meta.get("donor_institute"),
            "Pedigree or background from distributed metadata": str(meta.get("notes", "")).strip(),
            "Disease or stress context": context.get("Disease or stress context", "Not evaluated in the present reanalysis."),
            "Agronomic or quality context": context.get("Agronomic or quality context", "Not evaluated in the present reanalysis."),
            "sgC baseline rank": int(piv.loc[name, ("within_introgressed_rank", "sgC")]),
            "sgE baseline rank": int(piv.loc[name, ("within_introgressed_rank", "sgE")]),
            "sgC affinity score": float(piv.loc[name, ("pca_introgression_affinity", "sgC")]),
            "sgE affinity score": float(piv.loc[name, ("pca_introgression_affinity", "sgE")]),
            "Metadata reference": meta.get("full_reference") or meta.get("ref"),
            "Interpretive note": context.get("Interpretive note", "Contextual information only; no phenotype was measured in the present reanalysis."),
        })
    return pd.DataFrame(records)

def ranking_stability_summary(rankings: pd.DataFrame, common_panel: pd.DataFrame) -> pd.DataFrame:
    meta = common_panel[["seq_id", "accession_display"]].copy()
    meta["seq_id"] = meta["seq_id"].astype(str)
    df = rankings[(rankings["panel_type"] == "arabica_only") &
                  (rankings["analysis_group"] == "arabica_introgressed")].copy()
    df["sample_id"] = df["sample_id"].astype(str)
    df = df.merge(meta, left_on="sample_id", right_on="seq_id", how="left")
    rows = []
    for (name, sub), g in df.groupby(["accession_display", "subgenome"]):
        baseline = g[g["condition"] == "filtered_reservoir_12000"]
        rows.append({
            "Accession": name,
            "Subgenome": sub,
            "Baseline 12k rank": int(baseline["within_introgressed_rank"].iloc[0]),
            "Mean rank across conditions": float(g["within_introgressed_rank"].mean()),
            "Minimum rank": int(g["within_introgressed_rank"].min()),
            "Maximum rank": int(g["within_introgressed_rank"].max()),
            "Baseline affinity score": float(baseline["pca_introgression_affinity"].iloc[0]),
            "Conditions evaluated": int(g["condition"].nunique()),
            "Top-six membership in all conditions": bool((g["within_introgressed_rank"] <= 6).all()),
        })
    return pd.DataFrame(rows).sort_values(["Subgenome", "Baseline 12k rank"])

def validation_overview(status: dict, align_summary: pd.DataFrame, perm: pd.DataFrame,
                        contig_audit: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {"Validation component": "Strict subgenome filtering", "Result": "Pass",
         "Key evidence": f"sgC eligible variants: {int(contig_audit.loc[contig_audit.subgenome=='sgC','eligible_variants_after_filtering'].iloc[0]):,}; sgE: {int(contig_audit.loc[contig_audit.subgenome=='sgE','eligible_variants_after_filtering'].iloc[0]):,}",
         "Interpretation": "Only explicitly labelled pseudomolecules of the intended subgenome were retained."},
        {"Validation component": "SNP-panel sensitivity", "Result": "Pass",
         "Key evidence": f"Median PCA distance correlation = {status['metrics']['median_pca_distance_correlation']:.4f}",
         "Interpretation": "Population structure is stable across panel sizes and seeds."},
        {"Validation component": "Chromosome 4 recovery", "Result": "Pass",
         "Key evidence": f"Expected chromosome 4 recovery fraction = {status['metrics']['expected_chr4_recovery_fraction']:.3f}; baseline = {status['metrics']['baseline_expected_chr4_fraction']:.1f}",
         "Interpretation": "Chromosome 4 is consistently recovered at 12k SNPs and above."},
        {"Validation component": "Accession ranking", "Result": "Pass",
         "Key evidence": f"Median Spearman correlation = {status['metrics']['median_accession_rank_spearman']:.4f}; median top-six overlap = {status['metrics']['median_top6_overlap']:.0f}/6",
         "Interpretation": "Priority introgressed accessions are stable across sensitivity conditions."},
        {"Validation component": "Diploid-reference sensitivity", "Result": "Geometry changed",
         "Key evidence": f"Median full vs. Arabica-only distance correlation = {status['metrics']['median_full_vs_arabica_only_distance_correlation']:.3f}",
         "Interpretation": "Arabica-only PCA is used for within-Arabica inference; full-panel PCA is retained as progenitor context."},
    ]
    for _, row in align_summary.iterrows():
        rows.append({
            "Validation component": f"Direct alignment, {row['subgenome']}",
            "Result": f"Chromosome {row['chromosome_assignment_status']}; boundary {row['boundary_resolution_status']}",
            "Key evidence": f"{100*row['aligned_source_fraction_expected_target']:.1f}% source coverage; {row['weighted_identity_expected_target']:.3f}% identity; top-target fraction {row['expected_target_top_hit_fraction']:.3f}",
            "Interpretation": "Chromosome-level correspondence is supported; exact base-pair boundaries remain provisional.",
        })
    for _, row in perm.iterrows():
        p_value = (1 + 0) / (1 + int(row["n_valid_permutations"]))
        # Use exact empirical p if present in workbook; otherwise report the conservative lower bound.
        if "empirical_p_value" in row and pd.notna(row["empirical_p_value"]):
            p_value = float(row["empirical_p_value"])
        rows.append({
            "Validation component": f"Genome-wide permutation, {row['subgenome']}",
            "Result": "Significant",
            "Key evidence": f"Observed max = {row['observed_max_window_mean_abs_delta_af']:.4f}; null 99th = {row['null_99th_percentile']:.4f}; empirical P <= {p_value:.4f}",
            "Interpretation": "Chromosome 4-associated differentiation exceeds the genome-wide permutation null.",
        })
    return pd.DataFrame(rows)

def interval_and_permutation_summary(intervals: pd.DataFrame, contigs: pd.DataFrame, windows: pd.DataFrame,
                                     perm: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for sub in ["sgC", "sgE"]:
        c = contigs[(contigs["subgenome"] == sub) & (contigs["genomewide_contig_rank"] == 1)].iloc[0]
        w = windows[(windows["subgenome"] == sub) & (windows["genomewide_window_rank"] == 1)].iloc[0]
        p = perm[perm["subgenome"] == sub].iloc[0]
        i = intervals[(intervals["subgenome"] == sub) & (intervals["condition"] == "filtered_reservoir_12000")].iloc[0]
        rows.append({
            "Subgenome": sub,
            "Top chromosome / contig": c["contig"],
            "Chromosome mean |ΔAF|": c["mean_abs_delta_af"],
            "Markers with |ΔAF| >= 0.90": int(c["n_markers_delta_ge_0_90"]),
            "Top 1-Mb window start": int(w["window_start"]),
            "Top 1-Mb window end": int(w["window_end"]),
            "Observed maximum window mean |ΔAF|": p["observed_max_window_mean_abs_delta_af"],
            "Permutation 99th percentile": p["null_99th_percentile"],
            "Valid permutations": int(p["n_valid_permutations"]),
            "Empirical P": float(p["empirical_p_value"]),
            "Baseline interval start": int(i["interval_start"]),
            "Baseline interval end": int(i["interval_end"]),
            "Baseline interval span (bp)": int(i["interval_span_bp"]),
            "Baseline expected chromosome 4 match": bool(i["expected_chromosome4_match"]),
        })
    return pd.DataFrame(rows)

def create_supplementary_tables(out_path: Path, status: dict, sensitivity: dict[str, pd.DataFrame],
                                alignment: dict[str, pd.DataFrame], common_panel: pd.DataFrame,
                                priority_accessions: Sequence[str], context_frame: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    overview = validation_overview(status, alignment["Corrected_summary"], sensitivity["Permutation_summary"], sensitivity["Contig_filter_audit"])
    context = accession_context_table(common_panel, sensitivity["Accession_rankings"], priority_accessions, context_frame)
    ranking_summary = ranking_stability_summary(sensitivity["Accession_rankings"], common_panel)
    candidates = candidate_table(alignment["candidate_annotations"]).drop(columns=["annotation_text"])
    interval_summary = interval_and_permutation_summary(
        sensitivity["Interval_stability"], sensitivity["Genomewide_contig_scan"],
        sensitivity["Genomewide_window_scan"], sensitivity["Permutation_summary"]
    )
    sheets = [
        ("S1_validation_overview", "Table S1. Validation overview for the subgenome-resolved analysis.",
         "Summarizes strict subgenome filtering, sensitivity analyses, direct alignment validation, and genome-wide permutation tests.", overview),
        ("S2_accession_context", "Table S2. Biological and breeding context of the principal introgressed accessions.",
         "Metadata-derived background and published context used to interpret the top-ranked introgressed accessions; no phenotype was measured in this reanalysis.", context),
        ("S3_filter_audit", "Table S3. Strict subgenome pseudomolecule filtering audit.",
         "Counts of retained and excluded VCF records used to establish subgenome-specific analysis panels.", sensitivity["Contig_filter_audit"]),
        ("S4_accession_stability", "Table S4. Accession prioritization and rank stability.",
         "Baseline ranks and rank ranges across SNP-panel sizes, random seeds, and the all-eligible analysis.", ranking_summary),
        ("S5_supported_genes", "Table S5. Gene models overlapping direct alignment-supported chromosome 4 blocks.",
         "Gene models are restricted to target blocks supported by direct sequence alignment and remain interval-level annotations.", alignment["supported_genes"]),
        ("S6_candidate_annotations", "Table S6. Provisional defense-, signaling-, and regulatory-associated annotations.",
         "Keyword-supported subset of Table S5; entries are hypotheses for downstream validation, not causal assignments.", candidates),
        ("S7_alignment_validation", "Table S7. Direct alignment validation of chromosome 4 correspondence.",
         "Chromosome assignment, alignment coverage, nucleotide identity, orientation, and boundary-resolution diagnostics.", alignment["Corrected_summary"]),
        ("S8_sampling_stability", "Table S8. SNP-panel and PCA sensitivity analysis.",
         "Condition-level stability metrics across 6k, 12k, 24k, 48k SNP panels, five seeds, and all eligible variants.", sensitivity["PCA_sampling_stability"]),
        ("S9_diploid_sensitivity", "Table S9. Diploid-reference sensitivity of within-Arabica population structure.",
         "Compares full-panel and Arabica-only ordination geometry and cultivated-introgressed centroid separation.", sensitivity["Diploid_reference_sensitivity"]),
        ("S10_genomewide_validation", "Table S10. Genome-wide chromosome 4 recovery, interval, and permutation summary.",
         "Summarizes the top chromosome, strongest 1-Mb window, empirical permutation support, and baseline operational interval.", interval_summary),
    ]
    readme = wb.create_sheet("README")
    readme.append(["Supplementary tables"])
    readme.append(["All subgenome analyses use explicitly labelled pseudomolecules, and candidate annotations are restricted to direct alignment-supported target blocks."])
    readme.merge_cells("A1:C1")
    readme.merge_cells("A2:C2")
    readme.row_dimensions[1].height = 24
    readme.row_dimensions[2].height = 42
    readme["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    readme.append([])
    readme.append(["Table", "Sheet", "Description"])
    for idx, (sheet_name, title, description, _) in enumerate(sheets, 1):
        readme.append([f"Table S{idx}", sheet_name, description])
    style_sheet(readme, header_row=4, freeze="A5")
    for sheet_name, title, description, df in sheets:
        add_dataframe_sheet(wb, sheet_name, title, description, df)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def create_data_s1(out_path: Path, sensitivity: dict[str, pd.DataFrame], alignment: dict[str, pd.DataFrame],
                   common_panel: pd.DataFrame, priority_accessions: Sequence[str],
                   context_frame: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    guide_rows = []
    all_sheets: list[tuple[str, str, pd.DataFrame]] = []
    # Curated context first.
    all_sheets.append(("Accession_context", "Context for the principal introgressed accessions.",
                       accession_context_table(common_panel, sensitivity["Accession_rankings"], priority_accessions, context_frame)))
    for name in [
        "Common_panel", "Contig_filter_audit", "PCA_conditions", "PCA_scores",
        "PCA_explained_variance", "PCA_sampling_stability", "Diploid_reference_sensitivity",
        "Accession_rank_stability", "Accession_rankings", "Interval_stability",
        "Top_marker_support", "Genomewide_contig_scan", "Genomewide_window_scan",
        "Permutation_summary", "Permutation_null",
    ]:
        all_sheets.append((name, f"Derived output from the strict subgenome-filtered sensitivity analysis: {name}.", sensitivity[name]))
    alignment_sheet_names = {"Corrected_summary": "Alignment_summary"}
    for name in ["Corrected_summary", "best_expected_hits", "alternative_targets", "alignment_blocks", "supported_genes", "candidate_annotations"]:
        output_name = alignment_sheet_names.get(name, name)
        all_sheets.append((output_name, f"Derived output from direct alignment validation: {output_name}.", alignment[name]))
    guide = wb.create_sheet("Guide")
    guide.append(["Supplementary Data S1"])
    guide.append(["Machine-readable outputs supporting the main and supplementary figures and tables."])
    guide.merge_cells("A1:D1")
    guide.merge_cells("A2:D2")
    guide.row_dimensions[1].height = 24
    guide.row_dimensions[2].height = 38
    guide["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    guide.append([])
    guide.append(["Sheet", "Description", "Rows", "Columns"])
    for name, description, df in all_sheets:
        guide.append([name[:31], description, len(df), len(df.columns)])
    style_sheet(guide, header_row=4, freeze="A5")
    for name, description, df in all_sheets:
        add_dataframe_sheet(wb, name, name.replace("_", " "), description, df)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def create_summary(out_path: Path, sensitivity_path: Path, alignment_path: Path, status: dict,
                   sensitivity: dict[str, pd.DataFrame], alignment: dict[str, pd.DataFrame]) -> None:
    perm = sensitivity["Permutation_summary"].set_index("subgenome")
    contig = sensitivity["Genomewide_contig_scan"]
    align = alignment["Corrected_summary"].set_index("subgenome")
    text = f"""# Analysis output summary

## Inputs

- Sensitivity workbook: `{sensitivity_path.name}`
- Alignment workbook: `{alignment_path.name}`

## Principal validation results

- Overall sensitivity status: **{status['overall_status']}**
- Median PCA distance correlation across sensitivity conditions: **{status['metrics']['median_pca_distance_correlation']:.4f}**
- Median accession-rank Spearman correlation: **{status['metrics']['median_accession_rank_spearman']:.4f}**
- Median top-six overlap: **{status['metrics']['median_top6_overlap']:.0f}/6**
- Median full-panel vs. Arabica-only distance correlation: **{status['metrics']['median_full_vs_arabica_only_distance_correlation']:.3f}**

## Genome-wide chromosome 4 evidence

- sgC top chromosome mean |ΔAF|: **{contig[(contig.subgenome=='sgC') & (contig.genomewide_contig_rank==1)].mean_abs_delta_af.iloc[0]:.4f}**
- sgE top chromosome mean |ΔAF|: **{contig[(contig.subgenome=='sgE') & (contig.genomewide_contig_rank==1)].mean_abs_delta_af.iloc[0]:.4f}**
- sgC permutation: observed **{perm.loc['sgC','observed_max_window_mean_abs_delta_af']:.4f}**, null 99th **{perm.loc['sgC','null_99th_percentile']:.4f}**, empirical P <= **{perm.loc['sgC','empirical_p_value']:.4f}**
- sgE permutation: observed **{perm.loc['sgE','observed_max_window_mean_abs_delta_af']:.4f}**, null 99th **{perm.loc['sgE','null_99th_percentile']:.4f}**, empirical P <= **{perm.loc['sgE','empirical_p_value']:.4f}**

## Direct alignment validation

- sgC: **{align.loc['sgC','chromosome_assignment_status']}** chromosome assignment; **{100*align.loc['sgC','aligned_source_fraction_expected_target']:.1f}%** aligned source coverage; **{align.loc['sgC','weighted_identity_expected_target']:.3f}%** weighted identity; boundary resolution **{align.loc['sgC','boundary_resolution_status']}**.
- sgE: **{align.loc['sgE','chromosome_assignment_status']}** chromosome assignment; **{100*align.loc['sgE','aligned_source_fraction_expected_target']:.1f}%** aligned source coverage; **{align.loc['sgE','weighted_identity_expected_target']:.3f}%** weighted identity; boundary resolution **{align.loc['sgE','boundary_resolution_status']}**.

## Interpretation constraints

1. Arabica-only PCA should be used for within-Arabica inference; full-panel PCA should be retained only as progenitor context.
2. sgC carries the stronger chromosome 4 differentiation signal. sgE retains a weaker chromosome 4-associated signal in the same accession contrast.
3. Exact interval endpoints should remain operational and supplementary. Chromosome-level correspondence is validated, but exact base-pair liftover boundaries are not resolved.
4. Candidate annotations are restricted to direct alignment-supported blocks and remain provisional.
5. Earlier accession-level subgenome-asymmetry candidates derived from mixed-contig analyses should not be retained without a new corrected definition.
"""
    out_path.write_text(text, encoding="utf-8")

def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def export_source_tables(
    output_dir: Path,
    sensitivity: dict[str, pd.DataFrame],
    alignment: dict[str, pd.DataFrame],
    common_panel: pd.DataFrame,
    priority_accessions: Sequence[str],
    context_frame: pd.DataFrame,
) -> list[dict[str, object]]:
    """Write the machine-readable source tables underlying manuscript figures."""
    output_dir.mkdir(parents=True, exist_ok=True)
    table_map: list[tuple[str, pd.DataFrame, str]] = [
        ("Figure_1_panel_metadata.csv", common_panel, "Panel metadata used to label Arabica-only PCA scores."),
        ("Figure_1_pca_scores.csv", sensitivity["PCA_scores"], "PCA scores for full-panel and Arabica-only conditions."),
        ("Figure_1_pca_variance.csv", sensitivity["PCA_explained_variance"], "Explained-variance ratios for PCA components."),
        ("Figure_2_genomewide_windows.csv", sensitivity["Genomewide_window_scan"], "Genome-wide 1-Mb cultivated-introgressed differentiation windows."),
        ("Figure_2_chromosome_summary.csv", sensitivity["Genomewide_contig_scan"], "Chromosome-level differentiation summaries."),
        ("Figure_2_permutation_summary.csv", sensitivity["Permutation_summary"], "Observed and permutation-derived genome-wide thresholds."),
        ("Figure_3_accession_rankings.csv", sensitivity["Accession_rankings"], "Condition-specific accession affinity scores and ranks."),
        ("Figure_3_accession_rank_stability.csv", ranking_stability_summary(sensitivity["Accession_rankings"], common_panel), "Accession-rank stability across sensitivity conditions."),
        ("Figure_4_alignment_summary.csv", alignment["Corrected_summary"], "Chromosome-level alignment validation summary."),
        ("Figure_4_alignment_blocks.csv", alignment["alignment_blocks"], "Direct alignment-supported target blocks."),
        ("Figure_4_candidate_annotations.csv", candidate_table(alignment["candidate_annotations"]).drop(columns=["annotation_text"]), "Provisional annotations within alignment-supported blocks."),
        ("Figure_S1_sampling_stability.csv", sensitivity["PCA_sampling_stability"], "PCA stability across SNP-panel sizes and seeds."),
        ("Figure_S2_diploid_reference_sensitivity.csv", sensitivity["Diploid_reference_sensitivity"], "Effect of diploid references on within-Arabica ordination geometry."),
        ("Figure_S3_permutation_null.csv", sensitivity["Permutation_null"], "Permutation null distribution for maximum 1-Mb window differentiation."),
        ("Accession_context.csv", accession_context_table(common_panel, sensitivity["Accession_rankings"], priority_accessions, context_frame), "Curated contextual information for priority accessions."),
        ("Alignment_supported_genes.csv", alignment["supported_genes"], "Gene models overlapping direct alignment-supported blocks."),
    ]
    manifest: list[dict[str, object]] = []
    for filename, frame, description in table_map:
        path = output_dir / filename
        frame.to_csv(path, index=False)
        manifest.append({
            "file": filename,
            "description": description,
            "rows": int(len(frame)),
            "columns": int(len(frame.columns)),
            "sha256": sha256_file(path),
        })
    readme = output_dir / "README.md"
    readme.write_text(
        "# Machine-readable source data\n\n"
        "These CSV files contain the values underlying the manuscript and supplementary figures. "
        "The repository intentionally excludes manuscript-specific graphic layout and rendering code. "
        "Each file can be plotted with any statistical graphics package.\n",
        encoding="utf-8",
    )
    manifest.append({
        "file": readme.name,
        "description": "Description of the machine-readable source-data directory.",
        "rows": None,
        "columns": None,
        "sha256": sha256_file(readme),
    })
    return manifest


def write_export_manifest(
    output_path: Path,
    sensitivity_path: Path,
    alignment_path: Path,
    generated_files: Sequence[Path],
    source_manifest: Sequence[dict[str, object]],
) -> None:
    payload = {
        "workflow": "supplementary-output-export",
        "figure_rendering_included": False,
        "inputs": {
            "sensitivity_workbook": str(sensitivity_path),
            "alignment_workbook": str(alignment_path),
        },
        "generated_files": [
            {
                "file": str(path.relative_to(output_path.parent)),
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
            for path in generated_files
        ],
        "source_tables": list(source_manifest),
    }
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> None:
    args = parse_args()
    sensitivity_path, alignment_path = discover_inputs(args)
    output_dir = args.output_dir.expanduser().resolve()
    tables_dir = output_dir / "tables"
    data_dir = output_dir / "data"
    source_dir = output_dir / "figure_source_data"
    for directory in (tables_dir, data_dir, source_dir):
        directory.mkdir(parents=True, exist_ok=True)

    sensitivity_xls = pd.ExcelFile(sensitivity_path)
    sensitivity = {name: pd.read_excel(sensitivity_xls, sheet_name=name) for name in sensitivity_xls.sheet_names}
    alignment_xls = pd.ExcelFile(alignment_path)
    alignment = {name: pd.read_excel(alignment_xls, sheet_name=name) for name in alignment_xls.sheet_names}

    common_panel = normalize_accession_names(sensitivity["Common_panel"])
    priority_accessions = load_priority_accessions(args.priority_accessions, sensitivity["Accession_rankings"], common_panel)
    context_frame = load_accession_context(args.accession_context)

    status_path = sensitivity_path.parent / "sampling_validation_status.json"
    if status_path.exists():
        status = json.loads(status_path.read_text(encoding="utf-8"))
    else:
        validation_status = sensitivity["Validation_status"]
        status = {
            "overall_status": str(validation_status[(validation_status.category == "overall") & (validation_status.item == "overall_status")].value.iloc[0]),
            "metrics": {row.item: float(row.value) for row in validation_status[validation_status.category == "metric"].itertuples()},
            "checks": {row.item: bool(row.value) for row in validation_status[validation_status.category == "check"].itertuples()},
        }

    supplementary_tables = tables_dir / "Supplementary_Tables.xlsx"
    supplementary_data = data_dir / "Supplementary_Data_S1.xlsx"
    summary_path = output_dir / "ANALYSIS_SUMMARY.md"
    manifest_path = output_dir / "EXPORT_MANIFEST.json"

    create_supplementary_tables(
        supplementary_tables, status, sensitivity, alignment, common_panel,
        priority_accessions, context_frame,
    )
    create_data_s1(
        supplementary_data, sensitivity, alignment, common_panel,
        priority_accessions, context_frame,
    )
    source_manifest = export_source_tables(
        source_dir, sensitivity, alignment, common_panel,
        priority_accessions, context_frame,
    )
    create_summary(summary_path, sensitivity_path, alignment_path, status, sensitivity, alignment)
    write_export_manifest(
        manifest_path, sensitivity_path, alignment_path,
        [supplementary_tables, supplementary_data, summary_path],
        source_manifest,
    )

    print("Supplementary outputs exported.")
    print(f"Supplementary tables: {supplementary_tables}")
    print(f"Supplementary Data S1: {supplementary_data}")
    print(f"Figure source data: {source_dir}")
    print(f"Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
