#!/usr/bin/env python3
"""Shared utilities for the Coffea arabica revision-validation scripts."""
from __future__ import annotations

import gzip
import json
import math
import os
import re
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MASK64 = (1 << 64) - 1


def splitmix64(value: int) -> int:
    """Fast deterministic 64-bit mixing function."""
    z = (value + 0x9E3779B97F4A7C15) & MASK64
    z = ((z ^ (z >> 30)) * 0xBF58476D1CE4E5B9) & MASK64
    z = ((z ^ (z >> 27)) * 0x94D049BB133111EB) & MASK64
    return (z ^ (z >> 31)) & MASK64


def stable_text_hash(text: str) -> int:
    """Stable FNV-1a 64-bit hash, independent of Python hash randomization."""
    h = 0xCBF29CE484222325
    for byte in text.encode("utf-8", errors="replace"):
        h ^= byte
        h = (h * 0x100000001B3) & MASK64
    return h


def variant_priority_key(contig_hash: int, pos: int, ref: str, alt: str, seed: int) -> int:
    ref_alt = stable_text_hash(f"{ref}>{alt}")
    raw = contig_hash ^ ((int(pos) * 0xD6E8FEB86659FD93) & MASK64) ^ ref_alt ^ int(seed)
    return splitmix64(raw)


def normalize_token(value: object) -> Optional[str]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    token = str(value).strip().lower()
    if not token:
        return None
    token = re.sub(r"[^a-z0-9]+", "", token)
    return token or None


def normalize_species(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip()
    low = text.lower()
    if low == "coffea arabica":
        return "Coffea arabica"
    if low == "coffea canephora":
        return "Coffea canephora"
    if low == "coffea eugenioides":
        return "Coffea eugenioides"
    return text


def normalize_variety(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip()
    low = text.lower()
    if "introgress" in low:
        return "Introgressed"
    if "cultivated" in low:
        return "Cultivated"
    if "wild" in low:
        return "Wild"
    return text


def infer_analysis_group(species: object, variety: object) -> str:
    species_norm = normalize_species(species)
    variety_norm = normalize_variety(variety)
    if species_norm == "Coffea arabica":
        if variety_norm == "Introgressed":
            return "arabica_introgressed"
        if variety_norm == "Cultivated":
            return "arabica_cultivated"
        if variety_norm == "Wild":
            return "arabica_wild"
        return "arabica_other"
    if species_norm == "Coffea canephora":
        return "canephora"
    if species_norm == "Coffea eugenioides":
        return "eugenioides"
    return "other"


def load_metadata(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    rename_map = {
        "Seq.ID": "seq_id",
        "Collection location ": "collection_location",
        "species_name": "species_name",
        "variety(s)": "variety",
        "location_code(s)": "location_code",
        "country_of_origin(s)": "country_of_origin",
        "ploidy_level(s)": "ploidy_level",
        "Genome size (Gb)": "genome_size_gb",
        "genome_structure(s)": "genome_structure",
        "donor_institute(s)": "donor_institute",
        "notes(s)": "notes",
        "Full ref": "full_reference",
    }
    df = df.rename(columns=rename_map)
    required = ["seq_id", "accession_name", "species_name", "variety"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Metadata file is missing required columns: {missing}")
    df["seq_id"] = df["seq_id"].astype(str).str.strip()
    df["accession_name"] = df["accession_name"].fillna("").astype(str).str.strip()
    df["species_name"] = df["species_name"].map(normalize_species)
    df["variety"] = df["variety"].map(normalize_variety)
    df["analysis_group"] = [infer_analysis_group(s, v) for s, v in zip(df["species_name"], df["variety"])]
    return df


def build_metadata_lookup(metadata: pd.DataFrame) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for _, row in metadata.iterrows():
        canonical = str(row["seq_id"])
        for value in (row.get("seq_id"), row.get("accession_name")):
            token = normalize_token(value)
            if token:
                lookup[token] = canonical
    return lookup


def candidate_sample_names(sample_id: object) -> List[str]:
    value = str(sample_id).strip()
    candidates = [value]
    candidates.append(re.sub(r"_(sgc|sge)$", "", value, flags=re.IGNORECASE))
    candidates.append(re.sub(r"_eugenioides$", "", value, flags=re.IGNORECASE))
    candidates.append(re.sub(r"_canephora$", "", value, flags=re.IGNORECASE))
    candidates.append(value.replace("_sgC", "").replace("_sgE", ""))
    candidates.append(value.replace("_Eugenioides", "").replace("_Canephora", ""))
    cleaned: List[str] = []
    seen = set()
    for candidate in candidates:
        candidate = candidate.strip("_ ")
        if candidate and candidate not in seen:
            cleaned.append(candidate)
            seen.add(candidate)
    return cleaned


def reconcile_sample_id(sample_id: object, lookup: Mapping[str, str]) -> Tuple[Optional[str], str]:
    raw = str(sample_id)
    token = normalize_token(raw)
    if token and token in lookup:
        return lookup[token], "exact"
    for candidate in candidate_sample_names(raw):
        token = normalize_token(candidate)
        if token and token in lookup:
            return lookup[token], f"normalized:{candidate}"
    return None, "unresolved"


def find_file(root: Path, names: Sequence[str], required: bool = True) -> Optional[Path]:
    root = root.expanduser().resolve()
    for name in names:
        direct = root / name
        if direct.exists():
            return direct
    lower_names = {name.lower() for name in names}
    matches: List[Path] = []
    for path in root.rglob("*"):
        if path.is_file() and path.name.lower() in lower_names:
            matches.append(path)
    if matches:
        matches.sort(key=lambda p: (len(p.parts), str(p).lower()))
        return matches[0]
    if required:
        raise FileNotFoundError(f"Could not find any of {list(names)} under {root}")
    return None


def find_files_by_suffix(root: Path, suffixes: Sequence[str]) -> List[Path]:
    suffixes_lower = tuple(s.lower() for s in suffixes)
    return sorted(
        [p for p in root.rglob("*") if p.is_file() and p.name.lower().endswith(suffixes_lower)],
        key=lambda p: str(p).lower(),
    )


def read_vcf_header(path: Path) -> Tuple[List[str], Dict[str, Optional[int]], List[str]]:
    header_lines: List[str] = []
    contigs: Dict[str, Optional[int]] = {}
    samples: List[str] = []
    with gzip.open(path, "rt", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            if line.startswith("##"):
                header_lines.append(line.rstrip("\n"))
                if line.startswith("##contig="):
                    id_match = re.search(r"ID=([^,>]+)", line)
                    length_match = re.search(r"length=(\d+)", line)
                    if id_match:
                        contigs[id_match.group(1)] = int(length_match.group(1)) if length_match else None
                continue
            if line.startswith("#CHROM"):
                samples = line.rstrip("\n").split("\t")[9:]
                break
    if not samples:
        raise RuntimeError(f"No sample header found in {path}")
    return samples, contigs, header_lines


def parse_gt_dosage(sample_value: str, gt_index: int) -> int:
    fields = sample_value.split(":")
    if gt_index >= len(fields):
        return -1
    gt = fields[gt_index]
    if not gt or gt == "." or "." in gt:
        return -1
    alleles = re.split(r"[/|]", gt)
    if len(alleles) != 2:
        return -1
    try:
        values = [int(a) for a in alleles]
    except ValueError:
        return -1
    if any(a < 0 or a > 1 for a in values):
        return -1
    return int(sum(values))


def standardize_variant_rows(dosage_matrix: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """Mean-impute and standardize variants (rows); returns valid standardized rows and mask."""
    x = dosage_matrix.astype(np.float64, copy=True)
    x[x < 0] = np.nan
    with np.errstate(invalid="ignore"):
        means = np.nanmean(x, axis=1)
    valid_mean = ~np.isnan(means)
    if not valid_mean.any():
        return np.empty((0, x.shape[1]), dtype=np.float64), valid_mean
    x = x[valid_mean]
    means = means[valid_mean]
    nan_rows, nan_cols = np.where(np.isnan(x))
    if nan_rows.size:
        x[nan_rows, nan_cols] = means[nan_rows]
    stds = x.std(axis=1, ddof=1)
    variable = np.isfinite(stds) & (stds > 0)
    x = x[variable]
    means = means[variable]
    stds = stds[variable]
    if x.size == 0:
        final_mask = np.zeros_like(valid_mean, dtype=bool)
        return x, final_mask
    x = (x - means[:, None]) / stds[:, None]
    final_mask = np.zeros_like(valid_mean, dtype=bool)
    valid_indices = np.where(valid_mean)[0]
    final_mask[valid_indices[variable]] = True
    return x, final_mask


def pca_from_dosage(dosage_matrix: np.ndarray, n_components: int = 5) -> Tuple[np.ndarray, np.ndarray]:
    x_std, _ = standardize_variant_rows(dosage_matrix)
    if x_std.shape[0] == 0:
        raise ValueError("No variable variants available for PCA")
    x_samples = x_std.T
    u, s, _ = np.linalg.svd(x_samples, full_matrices=False)
    k = min(n_components, u.shape[1])
    scores = u[:, :k] * s[:k]
    eig = (s ** 2) / max(1, x_samples.shape[0] - 1)
    ratio = eig / eig.sum() if eig.sum() > 0 else np.zeros_like(eig)
    return scores, ratio[:k]


def pca_from_gram(gram: np.ndarray, n_variants: int, n_components: int = 5) -> Tuple[np.ndarray, np.ndarray]:
    if n_variants <= 0:
        raise ValueError("No variants accumulated for full-panel PCA")
    gram = (gram + gram.T) / 2.0
    eigvals, eigvecs = np.linalg.eigh(gram)
    order = np.argsort(eigvals)[::-1]
    eigvals = np.clip(eigvals[order], 0.0, None)
    eigvecs = eigvecs[:, order]
    k = min(n_components, eigvecs.shape[1])
    scores = eigvecs[:, :k] * np.sqrt(eigvals[:k])[None, :]
    ratio = eigvals / eigvals.sum() if eigvals.sum() > 0 else np.zeros_like(eigvals)
    return scores, ratio[:k]


def pairwise_distances(scores: np.ndarray) -> np.ndarray:
    diff = scores[:, None, :] - scores[None, :, :]
    return np.sqrt(np.sum(diff * diff, axis=2))


def upper_triangle_values(matrix: np.ndarray) -> np.ndarray:
    idx = np.triu_indices(matrix.shape[0], k=1)
    return matrix[idx]


def safe_pearson(x: np.ndarray, y: np.ndarray) -> float:
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = np.isfinite(x) & np.isfinite(y)
    if mask.sum() < 3:
        return float("nan")
    x = x[mask]
    y = y[mask]
    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")
    return float(np.corrcoef(x, y)[0, 1])


def rankdata_average(values: Sequence[float]) -> np.ndarray:
    return pd.Series(values, dtype=float).rank(method="average", na_option="keep").to_numpy(dtype=float)


def spearman_correlation(x: Sequence[float], y: Sequence[float]) -> float:
    return safe_pearson(rankdata_average(x), rankdata_average(y))


def procrustes_similarity(reference: np.ndarray, candidate: np.ndarray) -> Tuple[float, float]:
    """Orthogonal Procrustes similarity and normalized RMSE after centering/scaling."""
    x = np.asarray(reference, dtype=float)
    y = np.asarray(candidate, dtype=float)
    k = min(x.shape[1], y.shape[1])
    x = x[:, :k] - x[:, :k].mean(axis=0, keepdims=True)
    y = y[:, :k] - y[:, :k].mean(axis=0, keepdims=True)
    nx = np.linalg.norm(x)
    ny = np.linalg.norm(y)
    if nx == 0 or ny == 0:
        return float("nan"), float("nan")
    x /= nx
    y /= ny
    u, _, vt = np.linalg.svd(y.T @ x, full_matrices=False)
    rotation = u @ vt
    y_aligned = y @ rotation
    error = np.linalg.norm(x - y_aligned)
    similarity = max(0.0, 1.0 - error / math.sqrt(2.0))
    rmse = float(np.sqrt(np.mean((x - y_aligned) ** 2)))
    return float(similarity), rmse


def interval_jaccard(start_a: float, end_a: float, start_b: float, end_b: float) -> float:
    if not all(np.isfinite([start_a, end_a, start_b, end_b])):
        return float("nan")
    left = max(float(start_a), float(start_b))
    right = min(float(end_a), float(end_b))
    overlap = max(0.0, right - left)
    union = max(float(end_a), float(end_b)) - min(float(start_a), float(start_b))
    return overlap / union if union > 0 else float("nan")


def format_excel(path: Path, freeze: str = "A2") -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(bottom=Side(style="thin", color="BFBFBF"))
    for ws in wb.worksheets:
        ws.freeze_panes = freeze
        widths: Dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                if cell.value is not None:
                    widths[cell.column] = max(widths.get(cell.column, 0), min(55, len(str(cell.value)) + 2))
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = max(12, min(55, width))
    wb.save(path)


def write_excel(path: Path, sheets: Mapping[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe = re.sub(r"[\\/*?:\[\]]", "_", name)[:31]
            (df if df is not None else pd.DataFrame()).to_excel(writer, sheet_name=safe, index=False)
    format_excel(path)


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False, default=str)


def parse_int_list(value: str) -> List[int]:
    items = [int(x.strip()) for x in value.split(",") if x.strip()]
    if not items:
        raise ValueError("Expected at least one integer")
    return items


def chromosome_label_from_source_contig(contig: str, suffix: str) -> Optional[str]:
    match = re.search(r"chr_([A-Z])_sg_[CE]", str(contig), flags=re.IGNORECASE)
    if not match:
        return None
    number = ord(match.group(1).upper()) - ord("A") + 1
    return f"{number}{suffix.lower()}"


def chromosome_number_from_source_contig(contig: str) -> Optional[int]:
    match = re.search(r"chr_([A-Z])_sg_[CE]", str(contig), flags=re.IGNORECASE)
    if not match:
        return None
    return ord(match.group(1).upper()) - ord("A") + 1


def read_fasta_lengths(path: Path) -> Dict[str, int]:
    lengths: Dict[str, int] = {}
    current: Optional[str] = None
    count = 0
    opener = gzip.open if path.name.lower().endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            if line.startswith(">"):
                if current is not None:
                    lengths[current] = count
                current = line[1:].strip().split()[0]
                count = 0
            else:
                count += len(line.strip())
        if current is not None:
            lengths[current] = count
    return lengths


def extract_fasta_record(path: Path, record_id: str, output_path: Path) -> bool:
    opener = gzip.open if path.name.lower().endswith(".gz") else open
    found = False
    writing = False
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with opener(path, "rt", encoding="utf-8", errors="replace") as src, output_path.open("w", encoding="utf-8") as out:
        for line in src:
            if line.startswith(">"):
                current = line[1:].strip().split()[0]
                writing = current == record_id
                if writing:
                    found = True
                    out.write(line)
            elif writing:
                out.write(line)
    if not found:
        output_path.unlink(missing_ok=True)
    return found
